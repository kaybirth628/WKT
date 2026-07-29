"""WKT 新产品 BOM 表单 Excel 解析与批量导入。"""
from __future__ import annotations

import io
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from test_impl.order_management.cost_analysis.bom_form_import import (
    apply_batch_duplicate_part_hints,
    build_import_payload,
    parse_bom_workbook,
    preview_import_batch,
    preview_import_rows,
    revalidate_preview_item,
)
from test_impl.order_management.customer_name import (
    extract_customer_hint_from_filename,
    resolve_customer_from_hint,
)
from test_impl.order_management.cost_analysis.cost_store import CostStore
from test_impl.order_management.cost_analysis.record_service import CostRecordService
from test_impl.order_management.order_entry.line_store import LineStore

_TEST_SUPPLIER = "锦拓"


def _build_sample_bom_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "电机外壳"

    ws["A1"] = "昆山威可特精密电子有限公司 新产品BOM表"
    ws["A3"] = "客户"
    ws["B3"] = "精达"
    ws["D3"] = "产品名称"
    ws["E3"] = "电机外壳"
    ws["A4"] = "模具编号"
    ws["B4"] = "WKT-MJ-LV-25010"
    ws["D4"] = "产品料号"
    ws["E4"] = "EP1210.50.01"
    ws["A5"] = "模穴"
    ws["B5"] = "1*1"
    ws["D5"] = "产品单重"
    ws["E5"] = "253g"
    ws["G5"] = "材质"
    ws["H5"] = "ADC12"
    ws["J5"] = "机台&吨位"
    ws["K5"] = "280T"

    ws["A7"] = "产品制程"
    ws["A8"] = "工序1"
    ws["B8"] = "供应商"
    ws["C8"] = "工装"
    ws["D8"] = "工序2"
    ws["E8"] = "供应商"
    ws["F8"] = "工装"
    ws["G8"] = "工序3"
    ws["H8"] = "供应商"
    ws["A9"] = "压铸下料"
    ws["B9"] = "厂内"
    ws["C9"] = "压铸模一套"
    ws["D9"] = "抛丸"
    ws["E9"] = _TEST_SUPPLIER
    ws["G9"] = "车加工"
    ws["H9"] = "益佳波"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fill_sample_bom_sheet(ws, *, title: str, part_no: str, product_name: str) -> None:
    ws.title = title
    ws["A1"] = "昆山威可特精密电子有限公司 新产品BOM表"
    ws["A3"] = "客户"
    ws["B3"] = "精达"
    ws["D3"] = "产品名称"
    ws["E3"] = product_name
    ws["A4"] = "模具编号"
    ws["B4"] = "WKT-MJ-LV-25010"
    ws["D4"] = "产品料号"
    ws["E4"] = part_no
    ws["A5"] = "模穴"
    ws["B5"] = "1*1"
    ws["D5"] = "产品单重"
    ws["E5"] = "253g"
    ws["G5"] = "材质"
    ws["H5"] = "ADC12"
    ws["J5"] = "机台&吨位"
    ws["K5"] = "280T"
    ws["A7"] = "产品制程"
    ws["A8"] = "工序1"
    ws["B8"] = "供应商"
    ws["C8"] = "工装"
    ws["D8"] = "工序2"
    ws["E8"] = "供应商"
    ws["F8"] = "工装"
    ws["G8"] = "工序3"
    ws["H8"] = "供应商"
    ws["A9"] = "压铸下料"
    ws["B9"] = "厂内"
    ws["C9"] = "压铸模一套"
    ws["D9"] = "抛丸"
    ws["E9"] = _TEST_SUPPLIER
    ws["G9"] = "车加工"
    ws["H9"] = "益佳波"


def _build_duplicate_part_workbook() -> bytes:
    wb = Workbook()
    ws1 = wb.active
    _fill_sample_bom_sheet(ws1, title="电机外壳A", part_no="EP1210.50.01", product_name="电机外壳A")
    ws2 = wb.create_sheet("电机外壳B")
    _fill_sample_bom_sheet(ws2, title="电机外壳B", part_no="EP1210.50.01", product_name="电机外壳B")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class BomFormImportTests(unittest.TestCase):
    def test_parse_sample_workbook(self) -> None:
        rows = parse_bom_workbook(_build_sample_bom_workbook())
        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row["sheet_name"], "电机外壳")
        self.assertEqual(row["customer_name"], "精达")
        self.assertEqual(row["product_name"], "电机外壳")
        self.assertEqual(row["product_part_no"], "EP1210.50.01")
        self.assertEqual(row["unit_weight_g"], "253")
        self.assertEqual(row["material"], "ADC12")
        self.assertEqual(row["machine_tonnage"], "280T")
        codes = [p["code"] for p in row["processes"]]
        self.assertEqual(codes, ["01", "08", "12"])
        self.assertEqual(row["processes"][0]["supplier"], "")
        self.assertEqual(row["processes"][1]["supplier"], _TEST_SUPPLIER)

    def test_preview_tiers(self) -> None:
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        store = CostStore(db_path=":memory:")
        previews = preview_import_rows(parsed, store=store)
        self.assertEqual(len(previews), 1)
        self.assertEqual(previews[0]["tier"], "passed")

    def test_batch_duplicate_part_hints(self) -> None:
        parsed = parse_bom_workbook(_build_duplicate_part_workbook())
        store = CostStore(db_path=":memory:")
        previews = preview_import_rows(parsed, store=store)
        self.assertEqual(len(previews), 2)
        self.assertEqual(sum(1 for p in previews if p.get("duplicate_part_no")), 2)
        for preview in previews:
            self.assertTrue(
                any("本批料号重复" in i for i in preview.get("issues") or []),
                preview.get("issues"),
            )
            self.assertEqual(preview["tier"], "pending")
        cleared = [dict(p) for p in previews]
        for item in cleared:
            item["parsed"] = dict(item["parsed"])
            item["parsed"]["product_part_no"] = "UNIQUE-" + str(item["index"])
        self.assertEqual(apply_batch_duplicate_part_hints(cleared), 0)
        self.assertFalse(any(p.get("duplicate_part_no") for p in cleared))

    def test_sheet_customer_short_name_resolved_on_preview(self) -> None:
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        for row in parsed:
            row["customer_name"] = "大沃"
        store = CostStore(db_path=":memory:")
        batch = preview_import_batch(
            parsed,
            store=store,
            filename="大沃BOM.xlsx",
            customer_names=["苏州大沃工具科技有限公司", "苏州鑫福泰电子科技有限公司-东硕"],
        )
        self.assertEqual(batch["items"][0]["parsed"]["customer_name"], "苏州大沃工具科技有限公司")
        self.assertEqual(batch["customer_resolved"], "苏州大沃工具科技有限公司")
        self.assertEqual(batch["customer_error"], "")

    def test_import_overwrite_reupload_same_excel(self) -> None:
        """重复上传同一 Excel：料号相同则第二次覆盖第一次，不新增。"""
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        store = CostStore(db_path=":memory:")
        line_store = LineStore(db_path=":memory:")
        service = CostRecordService(store=store, line_store=line_store)
        payload = preview_import_rows(parsed, store=store)[0]["payload"]
        payload["customer_name"] = "苏州大沃工具科技有限公司"
        r1 = service.import_bom_rows([payload], skip_supplier_check=True, overwrite=True)
        r2 = service.import_bom_rows([payload], skip_supplier_check=True, overwrite=True)
        self.assertEqual(r1["created"], 1)
        self.assertEqual(r2["updated"], 1)
        self.assertEqual(r2["created"], 0)
        self.assertEqual(len(store.list_ids_by_part_no("EP1210.50.01")), 1)

    def test_import_overwrite_same_part_different_customer_alias(self) -> None:
        """旧记录客户为简称、新导入为全称时，仍按料号覆盖。"""
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        store = CostStore(db_path=":memory:")
        line_store = LineStore(db_path=":memory:")
        service = CostRecordService(store=store, line_store=line_store)
        old_payload = preview_import_rows(parsed, store=store)[0]["payload"]
        old_payload["customer_name"] = "大沃"
        service.import_bom_rows([old_payload], skip_supplier_check=True, overwrite=True)
        new_payload = dict(old_payload)
        new_payload["customer_name"] = "苏州大沃工具科技有限公司"
        new_payload["product_name"] = "电机外壳（第二次导入）"
        result = service.import_bom_rows([new_payload], skip_supplier_check=True, overwrite=True)
        self.assertEqual(result["updated"], 1)
        self.assertEqual(result["created"], 0)
        record = service.get_record(result["record_ids"][0])
        self.assertEqual(record.customer_name, "苏州大沃工具科技有限公司")
        self.assertEqual(record.product_name, "电机外壳（第二次导入）")

    def test_import_overwrite_same_part_no(self) -> None:
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        store = CostStore(db_path=":memory:")
        line_store = LineStore(db_path=":memory:")
        service = CostRecordService(store=store, line_store=line_store)
        payload = preview_import_rows(parsed, store=store)[0]["payload"]
        payload["customer_name"] = "苏州大沃工具科技有限公司"
        r1 = service.import_bom_rows([payload], skip_supplier_check=True, overwrite=True)
        self.assertEqual(r1["created"], 1)
        self.assertEqual(r1["updated"], 0)
        payload["product_name"] = "电机外壳（覆盖）"
        payload["machine_tonnage"] = "350T"
        r2 = service.import_bom_rows([payload], skip_supplier_check=True, overwrite=True)
        self.assertEqual(r2["updated"], 1)
        self.assertEqual(r2["created"], 0)
        self.assertEqual(len(store.list_ids_by_part_no("EP1210.50.01")), 1)
        record = service.get_record(r2["record_ids"][0])
        self.assertEqual(record.product_name, "电机外壳（覆盖）")
        self.assertEqual(record.machine_tonnage, "350T")

    def test_existing_part_preview_not_blocked(self) -> None:
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        store = CostStore(db_path=":memory:")
        line_store = LineStore(db_path=":memory:")
        service = CostRecordService(store=store, line_store=line_store)
        payload = preview_import_rows(parsed, store=store)[0]["payload"]
        payload["customer_name"] = "苏州大沃工具科技有限公司"
        service.import_bom_rows([payload], skip_supplier_check=True)
        previews = preview_import_rows(parsed, store=store)
        self.assertIn(previews[0]["tier"], ("passed", "pending"))
        self.assertFalse(
            any("覆盖" in i for i in previews[0]["issues"]),
            "解析预览不应查库提示覆盖，须完整保留各行供人工确认",
        )
        before_upload = revalidate_preview_item(
            previews[0]["parsed"],
            previews[0]["parsed"]["customer_name"],
            store=store,
            check_existing_db=True,
        )
        self.assertTrue(
            any("覆盖" in i for i in before_upload["issues"]),
            before_upload["issues"],
        )

    def test_revalidate_field_override_part_no(self) -> None:
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        store = CostStore(db_path=":memory:")
        preview = preview_import_rows(parsed, store=store)[0]
        updated = revalidate_preview_item(
            preview["parsed"],
            "苏州鑫福泰电子科技有限公司-东硕",
            store=store,
            fields={"product_part_no": "NEW-PART-001"},
        )
        self.assertEqual(updated["payload"]["product_part_no"], "NEW-PART-001")

    def test_revalidate_manual_customer_unblocks_row(self) -> None:
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        for row in parsed:
            row["customer_name"] = "DLS"
            row["sheet_customer_raw"] = "DLS"
            row["warnings"] = ["未找到客户「DLS」，请先在客商信息维护中建立客户档案后再导入 BOM"]
        store = CostStore(db_path=":memory:")
        blocked = preview_import_rows(parsed, store=store)[0]
        self.assertEqual(blocked["tier"], "blocked")
        fixed = revalidate_preview_item(
            blocked["parsed"],
            "苏州鑫福泰电子科技有限公司-东硕",
            store=store,
        )
        self.assertIn(fixed["tier"], ("passed", "pending"))
        self.assertEqual(
            fixed["payload"]["customer_name"],
            "苏州鑫福泰电子科技有限公司-东硕",
        )
        self.assertTrue(
            any("人工确认" in i for i in fixed["issues"]),
            fixed["issues"],
        )

    def test_build_import_payload_expands_customer_alias(self) -> None:
        payload = build_import_payload({"customer_name": "大沃", "processes": [{"code": "yz"}]})
        self.assertEqual(payload["customer_name"], "苏州大沃工具科技有限公司")

    def test_filename_dawo_product_bom_copy_suffix(self) -> None:
        self.assertEqual(
            extract_customer_hint_from_filename("大沃产品BOM(1)(1).xlsx"),
            "大沃",
        )
        resolved, err = resolve_customer_from_hint(
            "大沃",
            ["苏州大沃工具科技有限公司"],
        )
        self.assertEqual(resolved, "苏州大沃工具科技有限公司")
        self.assertEqual(err, "")

    def test_preview_meta_uses_sheet_customer_when_filename_fails(self) -> None:
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        for row in parsed:
            row["customer_name"] = "大沃"
        store = CostStore(db_path=":memory:")
        batch = preview_import_batch(
            parsed,
            store=store,
            filename="大沃产品BOM(1)(1).xlsx",
            customer_names=["苏州大沃工具科技有限公司"],
        )
        self.assertEqual(batch["customer_resolved"], "苏州大沃工具科技有限公司")
        self.assertEqual(batch["customer_error"], "")
        self.assertEqual(batch["items"][0]["parsed"]["customer_name"], "苏州大沃工具科技有限公司")

    def test_filename_customer_match(self) -> None:
        self.assertEqual(extract_customer_hint_from_filename("东硕BOM.xls"), "东硕")
        resolved, err = resolve_customer_from_hint(
            "东硕",
            ["苏州鑫福泰电子科技有限公司-东硕", "苏州大沃工具科技有限公司"],
        )
        self.assertEqual(resolved, "苏州鑫福泰电子科技有限公司-东硕")
        self.assertTrue(err == "" or "已匹配" in err)

    def test_filename_customer_missing_blocks(self) -> None:
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        for row in parsed:
            row["customer_name"] = ""
        store = CostStore(db_path=":memory:")
        batch = preview_import_batch(
            parsed,
            store=store,
            filename="不存在客户BOM.xls",
            customer_names=["苏州大沃工具科技有限公司"],
        )
        self.assertEqual(batch["customer_resolved"], "")
        self.assertIn("未找到客户", batch["customer_error"])
        self.assertEqual(batch["items"][0]["tier"], "blocked")

    def test_import_batch_zero_price(self) -> None:
        parsed = parse_bom_workbook(_build_sample_bom_workbook())
        store = CostStore(db_path=":memory:")
        line_store = LineStore(db_path=":memory:")
        service = CostRecordService(store=store, line_store=line_store)
        with patch(
            "test_impl.order_management.cost_analysis.record_service.list_profile_suppliers",
            return_value=[_TEST_SUPPLIER, "益佳波"],
        ):
            previews = preview_import_rows(parsed, store=store)
            result = service.import_bom_rows([previews[0]["payload"]], skip_supplier_check=True)
        self.assertEqual(result["imported"], 1)
        record = service.get_record(result["record_ids"][0])
        self.assertEqual(record.product_part_no, "EP1210.50.01")
        prices = {
            k: v for k, v in record.process_prices.items() if k != "__order__"
        }
        for entry in prices.values():
            if isinstance(entry, dict):
                self.assertEqual(entry.get("price"), "0")
            else:
                self.assertEqual(str(entry), "0")

    def test_parse_legacy_xls_workbook(self) -> None:
        try:
            import xlwt
        except ImportError:
            self.skipTest("xlwt not installed")
        wb = xlwt.Workbook()
        ws = wb.add_sheet("电机外壳")
        ws.write(2, 0, "客户")
        ws.write(2, 1, "精达")
        ws.write(3, 3, "产品料号")
        ws.write(3, 4, "EP1210.50.01")
        ws.write(2, 3, "产品名称")
        ws.write(2, 4, "电机外壳")
        ws.write(7, 0, "工序1")
        ws.write(7, 3, "工序2")
        ws.write(8, 0, "压铸下料")
        ws.write(8, 1, "厂内")
        ws.write(8, 3, "抛丸")
        ws.write(8, 4, _TEST_SUPPLIER)
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_bom_workbook(buf.getvalue(), filename="东硕BOM.xls")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["product_part_no"], "EP1210.50.01")

    def test_parse_vertical_process_layout(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "HUV9142V1"
        ws["A10"] = "制程:1"
        ws["B9"] = "工序"
        ws["C9"] = "可加工厂商"
        ws["B10"] = "压铸"
        ws["C10"] = "厂内"
        ws["A11"] = "制程:2"
        ws["B11"] = "冲切下料"
        ws["C11"] = "厂内"
        ws["A12"] = "制程:3"
        ws["B12"] = "振动研磨"
        ws["C12"] = "喜来来"
        ws["A13"] = "制程:4"
        ws["B13"] = "皮膜钝化（拉白）"
        ws["C13"] = "巧手"
        ws["A14"] = "制程:5"
        ws["B14"] = "全检出货"
        ws["C14"] = "厂内"
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_bom_workbook(buf.getvalue(), filename="东硕BOM.xlsx")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["product_part_no"], "/")
        self.assertEqual(rows[0]["product_name"], "HUV9142V1")
        codes = [p["code"] for p in rows[0]["processes"]]
        self.assertEqual(codes, ["01", "09", "15", "28"])


    def test_confirmed_process_aliases(self) -> None:
        cases = [
            ("清洗拉白", "17"),
            ("钝化拉白", "15"),
            ("拉白", "15"),
            ("铝挤", "32"),
            ("精冲下料", "01"),
            ("皮膜", "15"),
            ("攻牙", "11"),
        ]
        for raw, expected in cases:
            code, _, warnings = _resolve_process_name_for_test(raw)
            self.assertEqual(code, expected, raw)
            self.assertEqual(warnings, [], raw)

    def test_vibratory_deburr_expands_to_two_steps(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "TEST1"
        ws["A10"] = "制程:1"
        ws["B9"] = "工序"
        ws["C9"] = "供应商"
        ws["B10"] = "振动研磨去毛边"
        ws["C10"] = "喜来来"
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_bom_workbook(buf.getvalue())
        codes = [p["code"] for p in rows[0]["processes"]]
        self.assertEqual(codes, ["02", "09"])

    def test_skip_rivet_spring_process(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "TEST2"
        ws["A10"] = "制程:1"
        ws["B9"] = "工序"
        ws["B10"] = "铆合弹片"
        ws["A11"] = "制程:2"
        ws["B11"] = "全检"
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_bom_workbook(buf.getvalue())
        codes = [p["code"] for p in rows[0]["processes"]]
        self.assertEqual(codes, ["28"])

    def test_filename_customer_aliases(self) -> None:
        resolved, err = resolve_customer_from_hint(
            "欧菲光",
            ["安徽欧菲智能车联科技有限公司"],
        )
        self.assertEqual(resolved, "安徽欧菲智能车联科技有限公司")
        self.assertEqual(err, "")
        self.assertEqual(extract_customer_hint_from_filename("红黑BOM格式.xls"), "红黑")


    def test_vertical_bom_header_reads_part_no(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "HUV9142V1"
        ws["A3"] = "客戶"
        ws["C3"] = "产品名称"
        ws["D3"] = "模具编号"
        ws["E3"] = "产品料号"
        ws["A4"] = "东硕"
        ws["C4"] = "HUV9142V1散热片"
        ws["D4"] = "客户移模"
        ws["E4"] = "R08AA15A15A80000-X"
        ws["B5"] = "模穴"
        ws["C5"] = "产品单重(G)"
        ws["A6"] = "1*2"
        ws["C6"] = "126"
        ws["A10"] = "制程:1"
        ws["B9"] = "工序"
        ws["B10"] = "压铸"
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_bom_workbook(buf.getvalue())
        self.assertEqual(rows[0]["product_part_no"], "R08AA15A15A80000-X")
        self.assertEqual(rows[0]["product_name"], "HUV9142V1散热片")
        self.assertEqual(rows[0]["cavity"], "1*2")
        self.assertEqual(rows[0]["unit_weight_g"], "126")

    def test_part_no_rejects_quantity_label(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "HUV9142V1"
        ws["E3"] = "产品料号"
        ws["E4"] = "R08AA15A15A80000-X"
        ws["F20"] = "数量"
        ws["A10"] = "制程:1"
        ws["B9"] = "工序"
        ws["B10"] = "压铸"
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_bom_workbook(buf.getvalue())
        self.assertEqual(rows[0]["product_part_no"], "R08AA15A15A80000-X")


    def test_split_dual_product_sheet(self) -> None:
        rows = parse_bom_workbook(
            open("data/bom_import_audit/怡利BOM.xls", "rb").read(),
            filename="怡利BOM.xls",
        )
        metal = [r for r in rows if r["sheet_name"] == "金属转轴"]
        self.assertEqual(len(metal), 2)
        self.assertEqual(metal[0]["product_part_no"], "1A104D0A001-00")
        self.assertEqual(metal[1]["product_part_no"], "1A104D0A003-00")
        self.assertEqual(metal[0]["product_name"], "EL-104D-0003金属转轴A")
        self.assertEqual(metal[1]["product_name"], "EL-104D-0003金属转轴B")
        self.assertEqual(metal[0]["unit_weight_g"], "39.6")
        self.assertEqual(metal[1]["unit_weight_g"], "13.2")


    def test_unfilled_part_no_shows_slash_and_imports(self) -> None:
        path = Path("Demo/BOM/锐霸产品BOM.xls")
        if not path.is_file():
            self.skipTest("missing Demo/BOM/锐霸产品BOM.xls")
        rows = parse_bom_workbook(path.read_bytes(), filename=path.name)
        slash_rows = [r for r in rows if r["product_part_no"] == "/"]
        self.assertGreaterEqual(len(slash_rows), 5)
        store = CostStore(db_path=":memory:")
        previews = preview_import_rows(slash_rows[:3], store=store)
        for p in previews:
            self.assertIn(p["tier"], ("passed", "pending"))
            self.assertNotEqual(p["tier"], "blocked")
        from test_impl.order_management.cost_analysis.record_service import CostRecordService
        from test_impl.order_management.order_entry.line_store import LineStore

        service = CostRecordService(store=store, line_store=LineStore(db_path=":memory:"))
        result = service.import_bom_rows(
            [p["payload"] for p in previews],
            skip_supplier_check=True,
        )
        self.assertEqual(result["imported"], 3)
        self.assertEqual(result["created"], 3)

    def test_ruiba_sheet_count_matches_bom_rows(self) -> None:
        """锐霸：46 Sheet → 46 条 BOM（1 Sheet 1 BOM）。"""
        path = Path("Demo/BOM/锐霸产品BOM.xls")
        if not path.is_file():
            self.skipTest("missing Demo/BOM/锐霸产品BOM.xls")
        import xlrd

        raw = path.read_bytes()
        book = xlrd.open_workbook(file_contents=raw)
        sheet_count = len(book.sheets())
        rows = parse_bom_workbook(raw, filename=path.name)
        self.assertEqual(sheet_count, len(rows), (sheet_count, len(rows)))
        by_sheet = {}
        for r in rows:
            by_sheet.setdefault(r["sheet_name"], 0)
            by_sheet[r["sheet_name"]] += 1
        multi = {s: c for s, c in by_sheet.items() if c > 1}
        self.assertEqual(multi, {}, f"同 Sheet 不应拆多行: {multi}")

    def test_part_no_slash_preserved_ruiba(self) -> None:
        """锐霸：11*000000/08016-01 为完整料号，不可按 / 拆成两条。"""
        path = Path("Demo/BOM/锐霸产品BOM.xls")
        if not path.is_file():
            self.skipTest("missing Demo/BOM/锐霸产品BOM.xls")
        rows = parse_bom_workbook(path.read_bytes(), filename=path.name)
        by_sheet = {r["sheet_name"]: r for r in rows}
        self.assertEqual(len([r for r in rows if r["sheet_name"] == "819 头壳"]), 1)
        self.assertEqual(by_sheet["819 头壳"]["product_part_no"], "11*000000/08016-01")
        self.assertEqual(by_sheet["826头壳"]["product_part_no"], "11*000000/09016-01")
        self.assertEqual(by_sheet["826A头壳"]["product_part_no"], "11*000000/10016-01")

    def test_one_sheet_one_bom_shared_header(self) -> None:
        """多 Sheet 共用表头多型号品名 + 相同客户料号 → 按 Sheet 拆成多条 BOM。"""
        wb = Workbook()
        wb.remove(wb.active)
        combined = "SD-819/826/826A头壳"
        for title in ("SD-819", "SD-826", "SD-826A"):
            ws = wb.create_sheet(title=title)
            ws["D3"] = "产品名称"
            ws["E3"] = combined
            ws["D4"] = "产品料号"
            ws["E4"] = 11000061
            ws["A8"] = "工序1"
            ws["A9"] = "压铸"
            ws["B9"] = "厂内"
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_bom_workbook(buf.getvalue(), filename="test.xlsx")
        self.assertEqual(len(rows), 3)
        by_sheet = {r["sheet_name"]: r for r in rows}
        self.assertEqual(by_sheet["SD-819"]["product_part_no"], "11000061")
        self.assertEqual(by_sheet["SD-826"]["product_part_no"], "11000061")
        self.assertEqual(by_sheet["SD-826A"]["product_part_no"], "11000061")
        self.assertEqual(by_sheet["SD-819"]["product_name"], "SD-819头壳")
        self.assertEqual(by_sheet["SD-826"]["product_name"], "SD-826头壳")
        self.assertEqual(by_sheet["SD-826A"]["product_name"], "SD-826A头壳")
        dup_warnings = [
            w
            for r in rows
            for w in (r.get("warnings") or [])
            if "重复" in w
        ]
        self.assertEqual(len(dup_warnings), 2)

    def test_coerce_excel_float_part_no(self) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "TST-PART"
        ws["D4"] = "产品料号"
        ws["E4"] = 11000061.0
        ws["A8"] = "工序1"
        ws["A9"] = "压铸"
        buf = io.BytesIO()
        wb.save(buf)
        rows = parse_bom_workbook(buf.getvalue())
        self.assertEqual(rows[0]["product_part_no"], "11000061")


def _resolve_process_name_for_test(raw: str):
    from test_impl.order_management.cost_analysis import bom_form_import as mod

    return mod._resolve_process_name(raw)


if __name__ == "__main__":
    unittest.main()
