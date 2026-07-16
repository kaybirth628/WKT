import io
import unittest
from decimal import Decimal

from test_impl.order_management.delivery_note.wkt_document import (
    WktDeliveryDocument,
    WktDeliveryLine,
)
from test_impl.order_management.delivery_note.wkt_xlsx import build_xlsx_bytes

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class TestWktXlsx(unittest.TestCase):
    def test_meta_uses_merged_cells_not_single_column(self) -> None:
        if load_workbook is None:
            self.skipTest("openpyxl not installed")
        doc = WktDeliveryDocument(
            title_company="苏州允锐液压设备有限公司",
            doc_no="WKT202607150004",
            ship_date_cn="2026年7月15日",
            receiver_company="苏州允锐液压设备有限公司",
            receiver_address="江苏省苏州市工业园区唯亭镇唯新路6号厂房",
            receiver_contact="肖丽红",
            supplier_name="昆山威可特精密电子有限公司",
            supplier_address="江苏省昆山市锦溪镇锦昌路15号",
            supplier_phone="0512-50152121",
            lines=[
                WktDeliveryLine(
                    order_no="CG20260527049",
                    customer_part_no="06.2CPV4N.011",
                    product_name="PV4-N-011/手柄夹块-下",
                    spec="ADC12",
                    unit="pcs",
                    qty="1",
                ),
                WktDeliveryLine(
                    order_no="CG20260527049",
                    customer_part_no="06.2CPV4N.010",
                    product_name="PV4-N-010/手柄夹块-上",
                    spec="ADC12",
                    unit="pcs",
                    qty="1",
                ),
            ],
            total_qty="2",
            footer_note="一式四联：白联存根、红联仓库、蓝联、黄联IQC",
        )
        data = build_xlsx_bytes(doc)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual(ws["B4"].value, "WKT202607150004")
        self.assertEqual(ws["G4"].value, "2026年7月15日")
        self.assertEqual(ws["B6"].value, doc.receiver_address)
        self.assertEqual(ws["A4"].value, "送货单号：")
        self.assertNotIn(doc.doc_no, str(ws["A4"].value or ""))
        merges = {str(m) for m in ws.merged_cells.ranges}
        self.assertIn("B4:E4", merges)
        self.assertIn("G4:I4", merges)
        self.assertIn("B6:E6", merges)
        self.assertEqual(ws.cell(10, 6).value, "1")
        self.assertEqual(ws.cell(11, 6).value, "1")
        self.assertEqual(ws.cell(13, 6).value, "2")


if __name__ == "__main__":
    unittest.main()
