import os
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from test_impl.order_management.delivery_note import DeliveryNoteService, DeliveryTemplateStore
from test_impl.order_management.order_entry import OrderLineService

from bom_helpers import seed_bom

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None  # type: ignore

_TEST_SUPPLIER = "测试CNC厂"


class TestDeliveryNote(unittest.TestCase):
    def setUp(self) -> None:
        self.tmp = tempfile.TemporaryDirectory()
        self.tpl_root = Path(self.tmp.name) / "delivery_templates"
        self.svc = OrderLineService(db_path=":memory:")
        self.templates = DeliveryTemplateStore(self.tpl_root)
        self.dn = DeliveryNoteService(self.svc, self.templates)
        self.supplier_patcher = patch(
            "test_impl.order_management.cost_analysis.record_service.list_profile_suppliers",
            return_value=[_TEST_SUPPLIER],
        )
        self.supplier_patcher.start()

    def tearDown(self) -> None:
        self.supplier_patcher.stop()
        self.tmp.cleanup()

    def _seed_bom(self, *, customer_name: str, product_part_no: str, product_name: str = "") -> None:
        seed_bom(
            self.svc._bom._records,
            customer_name=customer_name,
            product_part_no=product_part_no,
            product_name=product_name or product_part_no,
        )

    def test_builtin_context_after_ship(self) -> None:
        line = self.svc.create_line(
            {
                "customer": "怡利",
                "order_date": "2026-05-26",
                "order_no": "PO-DN-1",
                "product_spec": "散热片",
                "po_qty": "100",
                "shipped_qty": "0",
                "unit": "PCS",
            }
        )
        _, ev = self.svc.ship_line(line.id, "40")
        ctx = self.dn.build_context(ev)
        self.assertEqual(ctx["客户"], "怡利")
        self.assertIn(ctx["本次出货"], ("40", "40.0"))
        self.assertEqual(ctx["订单号"], "PO-DN-1")

    def test_wkt_standard_xlsx_after_ship(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl not installed")
        self._seed_bom(customer_name="爱毕黎", product_part_no="P-001", product_name="双嘴钳")
        line = self.svc.create_line(
            {
                "customer": "爱毕黎",
                "order_no": "PO-1",
                "product_spec": "双嘴钳",
                "customer_part_no": "P-001",
                "material": "钢",
                "po_qty": "100",
                "unit": "pcs",
            }
        )
        _, ev = self.svc.ship_line(line.id, "51")
        doc = self.dn.build_wkt_document(ev.id)
        self.assertEqual(doc.title_company, "爱毕黎")
        self.assertEqual(doc.lines[0].qty, "51")
        self.assertIn("威可特", doc.supplier_name)
        kind, payload = self.dn.render_for_event(ev.id)
        self.assertEqual(kind, "xlsx")
        data, _fname = payload
        self.assertTrue(len(data) > 500)
        import io

        from openpyxl import load_workbook

        wb2 = load_workbook(io.BytesIO(data))
        self.assertEqual(wb2.active["A1"].value, "爱毕黎")
        self.assertEqual(wb2.active["A2"].value, "送货单")

    def test_custom_excel_template_for_mapped_customer(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl not installed")
        files_dir = self.tpl_root / "files"
        files_dir.mkdir(parents=True, exist_ok=True)
        tpl_name = "专用客.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "{{客户}}"
        ws["A2"] = "{{订单号}}"
        ws["A3"] = "{{本次出货}}"
        wb.save(files_dir / tpl_name)
        self.templates.set_customer_template("专用客", tpl_name)

        line = self.svc.create_line(
            {
                "customer": "专用客",
                "order_no": "PO-S",
                "product_spec": "测试件",
                "po_qty": "10",
                "unit": "PCS",
            }
        )
        _, ev = self.svc.ship_line(line.id, "3")
        kind, payload = self.dn.render_for_event(ev.id)
        self.assertEqual(kind, "xlsx")
        data, _fname = payload
        from openpyxl import load_workbook
        import io

        wb2 = load_workbook(io.BytesIO(data))
        self.assertEqual(wb2.active["A1"].value, "专用客")
        self.assertEqual(wb2.active["A2"].value, "PO-S")

    def test_ship_ui_mode_custom(self) -> None:
        files_dir = self.tpl_root / "files"
        files_dir.mkdir(parents=True, exist_ok=True)
        tpl_name = "专用客.xlsx"
        (files_dir / tpl_name).write_bytes(b"fake")
        self.templates.set_customer_template("专用客", tpl_name)
        ui = self.dn.ship_ui_mode("专用客")
        self.assertEqual(ui["mode"], "custom_excel")
        self.assertIn("/api/delivery-templates/raw", ui["raw_download_url"])

    def test_open_custom_excel_local(self) -> None:
        from unittest.mock import patch

        if Workbook is None:
            self.skipTest("openpyxl not installed")
        files_dir = self.tpl_root / "files"
        files_dir.mkdir(parents=True, exist_ok=True)
        tpl_name = "本地客.xlsx"
        wb = Workbook()
        wb.active["A1"] = "{{客户}}"
        wb.save(files_dir / tpl_name)
        self.templates.set_customer_template("本地客", tpl_name)
        line = self.svc.create_line(
            {"customer": "本地客", "order_no": "PO-L", "product_spec": "件", "po_qty": "1"}
        )
        _, ev = self.svc.ship_line(line.id, "1")
        with patch("test_impl.order_management.delivery_note.custom_excel_attachment.open_in_excel"):
            out = self.dn.open_custom_excel_local(ev.id)
        self.assertTrue(out.get("ok"))
        self.assertTrue(out.get("auto_filled"))
        rel = self.svc._store.get_shipment_attachment(ev.id)
        self.assertTrue(rel.endswith(".xlsx"))

    def test_default_wkt_for_unmapped_customer(self) -> None:
        meta = self.dn.template_info("普通客户")
        self.assertTrue(meta["is_wkt_standard"])

    def test_ship_with_delivery_note_snapshot(self) -> None:
        import json

        line = self.svc.create_line(
            {
                "customer": "确认客",
                "order_no": "PO-C",
                "product_spec": "垫片",
                "po_qty": "50",
                "unit": "pcs",
            }
        )
        note = {
            "receiver_company": "确认客有限公司",
            "receiver_address": "测试地址",
            "lines": [{"qty": "10", "batch_no": "B001", "product_name": "垫片"}],
            "warehouse_manager": "张三",
        }
        _, ev = self.svc.ship_line(line.id, "10", delivery_note=note)
        raw = self.svc._store.get_shipment_delivery_note_json(ev.id)
        snap = json.loads(raw)
        self.assertEqual(snap["receiver_address"], "测试地址")
        self.assertEqual(snap["lines"][0]["batch_no"], "B001")
        doc = self.dn.build_wkt_document(ev.id)
        self.assertEqual(doc.warehouse_manager, "张三")

    def test_jinmai_custom_excel_uses_wkt_doc_prefix(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl not installed")
        customer = "上海金脉电子科技有限公司"
        files_dir = self.tpl_root / "files"
        files_dir.mkdir(parents=True, exist_ok=True)
        tpl_name = "金脉.xlsx"
        wb = Workbook()
        wb.active["A1"] = "{{送货单号}}"
        wb.save(files_dir / tpl_name)
        self.templates.set_customer_template(customer, tpl_name)
        line = self.svc.create_line(
            {
                "customer": customer,
                "order_no": "PO-JM",
                "product_spec": "测试件",
                "po_qty": "5",
                "unit": "PCS",
            }
        )
        _, ev = self.svc.ship_line(line.id, "2")
        ctx = self.dn.build_context(ev)
        self.assertTrue(str(ctx.get("送货单号") or "").startswith("WKT"))

    def test_resolve_doc_no_ignores_legacy_jm_prefix(self) -> None:
        import json

        customer = "上海金脉电子科技有限公司"
        line = self.svc.create_line(
            {
                "customer": customer,
                "order_no": "PO-LEG",
                "product_spec": "测试件",
                "po_qty": "1",
                "unit": "PCS",
            }
        )
        _, ev = self.svc.ship_line(line.id, "1")
        self.svc._store.save_shipment_delivery_note(
            ev.id,
            json.dumps({"doc_no": "JM202606230025"}, ensure_ascii=False),
        )
        doc_no = self.dn.resolve_delivery_doc_no(ev.id)
        self.assertTrue(doc_no.startswith("WKT"))
        self.assertNotEqual(doc_no, "JM202606230025")

    def test_batch_custom_excel_expands_detail_rows(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl not installed")
        import io

        from openpyxl import load_workbook

        customer = "合并客"
        files_dir = self.tpl_root / "files"
        files_dir.mkdir(parents=True, exist_ok=True)
        tpl_name = "batch.xlsx"
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "{{送货单号}}"
        ws["A2"] = "{{序号}}"
        ws["B2"] = "{{客户料号}}"
        ws["C2"] = "{{订单号}}"
        ws["D2"] = "{{本次出货}}"
        ws["A3"] = "合计"
        ws["D3"] = "{{合计}}"
        wb.save(files_dir / tpl_name)
        self.templates.set_customer_template(customer, tpl_name)

        self._seed_bom(customer_name=customer, product_part_no="P-A", product_name="件A")
        self._seed_bom(customer_name=customer, product_part_no="P-B", product_name="件B")
        line_a = self.svc.create_line(
            {
                "customer": customer,
                "order_no": "PO-A",
                "customer_part_no": "P-A",
                "product_spec": "件A",
                "po_qty": "10",
                "unit": "PCS",
            }
        )
        line_b = self.svc.create_line(
            {
                "customer": customer,
                "order_no": "PO-B",
                "customer_part_no": "P-B",
                "product_spec": "件B",
                "po_qty": "10",
                "unit": "PCS",
            }
        )
        _, ev_a = self.svc.ship_line(line_a.id, "1")
        _, ev_b = self.svc.ship_line(line_b.id, "2")
        header, line_ctxs = self.dn.build_batch_fill_contexts([ev_a.id, ev_b.id])
        data = self.dn.fill_excel_bytes(self.templates.resolve_template_path(customer), header, line_ctxs)
        out = load_workbook(io.BytesIO(data)).active
        self.assertEqual(out["C2"].value, "PO-A")
        self.assertEqual(out["C3"].value, "PO-B")
        self.assertEqual(out["B2"].value, "P-A")
        self.assertEqual(out["B3"].value, "P-B")
        self.assertEqual(out["D4"].value, header["合计"])

    def test_batch_bilingual_template_keeps_total_row(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl not installed")
        import io

        from openpyxl import load_workbook

        from test_impl.order_management.delivery_note.bilingual_template import save_bilingual_template

        customer = "上海金脉电子科技有限公司"
        files_dir = self.tpl_root / "files"
        files_dir.mkdir(parents=True, exist_ok=True)
        tpl_name = "jinmai_batch.xlsx"
        save_bilingual_template(customer, files_dir / tpl_name)
        self.templates.set_customer_template(customer, tpl_name)

        self._seed_bom(customer_name=customer, product_part_no="P-A", product_name="件A")
        self._seed_bom(customer_name=customer, product_part_no="P-B", product_name="件B")
        line_a = self.svc.create_line(
            {
                "customer": customer,
                "order_no": "PO-A",
                "customer_part_no": "P-A",
                "product_spec": "件A",
                "po_qty": "10",
                "unit": "PCS",
            }
        )
        line_b = self.svc.create_line(
            {
                "customer": customer,
                "order_no": "PO-B",
                "customer_part_no": "P-B",
                "product_spec": "件B",
                "po_qty": "10",
                "unit": "PCS",
            }
        )
        _, ev_a = self.svc.ship_line(line_a.id, "1")
        _, ev_b = self.svc.ship_line(line_b.id, "2")
        header, line_ctxs = self.dn.build_batch_fill_contexts([ev_a.id, ev_b.id])
        data = self.dn.fill_excel_bytes(self.templates.resolve_template_path(customer), header, line_ctxs)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertEqual(ws["B13"].value, "P-A")
        self.assertEqual(ws["B14"].value, "P-B")
        self.assertEqual(ws["A15"].value, "合计 Total")
        self.assertEqual(ws["F15"].value, header["合计"])
        self.assertEqual(header["合计"], "3")
        for row_idx in (13, 14):
            for col in range(1, 8):
                border = ws.cell(row_idx, col).border
                self.assertEqual(border.left.style, "thin")
                self.assertEqual(border.right.style, "thin")
        for col in (1, 6):
            border = ws.cell(15, col).border
            self.assertEqual(border.left.style, "thin")
            self.assertEqual(border.right.style, "thin")
        self.assertTrue(any(str(m) == "A16:G16" for m in ws.merged_cells.ranges))
        self.assertEqual(ws.cell(16, 1).border.left.style, "thin")

    def test_xunbo_custom_placeholders_filled(self) -> None:
        if Workbook is None:
            self.skipTest("openpyxl not installed")
        import io

        from openpyxl import load_workbook

        customer = "迅铂科技（常州）有限公司"
        tpl_path = self.tpl_root / "files" / "迅铂送货单.xlsx"
        tpl_path.parent.mkdir(parents=True, exist_ok=True)
        if not tpl_path.is_file():
            src = Path(__file__).resolve().parents[1] / "data" / "delivery_templates" / "files" / "迅铂送货单.xlsx"
            if src.is_file():
                tpl_path.write_bytes(src.read_bytes())
            else:
                wb = Workbook()
                ws = wb.active
                ws["B7"] = "{{订单号}}"
                wb.save(tpl_path)
        wb = load_workbook(tpl_path)
        ws = wb.active
        ws["B2"] = "{{送货地点}}"
        ws["B3"] = "{{订单下发抬头}}"
        wb.save(tpl_path)
        self.templates.set_customer_template(customer, "迅铂送货单.xlsx")
        self._seed_bom(customer_name=customer, product_part_no="P-XB", product_name="测试件")
        line = self.svc.create_line(
            {
                "customer": customer,
                "order_no": "PO-XB",
                "customer_part_no": "P-XB",
                "product_spec": "测试件",
                "po_qty": "5",
                "unit": "PCS",
            }
        )
        _, ev = self.svc.ship_line(line.id, "2")
        ctx = self.dn.build_context(ev)
        data = self.dn.fill_excel_bytes(tpl_path, ctx)
        ws = load_workbook(io.BytesIO(data)).active
        self.assertIn("无锡", str(ws["B2"].value or ""))
        self.assertIn("迅铂", str(ws["B3"].value or ""))
        self.assertEqual(ctx.get("送货地点"), ctx.get("收货地址"))
        self.assertEqual(ctx.get("订单下发抬头"), ctx.get("收货公司"))


if __name__ == "__main__":
    unittest.main()
