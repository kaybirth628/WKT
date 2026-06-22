"""双语送货单（金脉/金棒等客户专用）Excel 版式生成。"""
from __future__ import annotations

from pathlib import Path
from typing import Dict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
except ImportError:
    Workbook = None  # type: ignore

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 客户抬头与页脚（可按客户扩展）
CUSTOMER_LAYOUT: Dict[str, dict] = {
    "浙江金棒运动器材有限公司": {
        "brand_lines": (
            "浙江金棒",
            "ZHEJIANG JINBANG SPORTS EQUIPMENT CO., LTD.",
        ),
        "footer_note": "（一式2份，金棒一份，如需要供应商可带回一份）",
        "doc_prefix_default": "JB",
        "material_label": "金棒物料号/供应商料号",
    },
    "上海金脉电子科技有限公司": {
        "brand_lines": (
            "G-PULSE 金脉",
            "SHANGHAI G-PULSE ELECTRONICS TECHNOLOGY CO., LTD.",
        ),
        "footer_note": "（一式2份，金脉一份，如需要供应商可带回一份）",
        "doc_prefix_default": "JM",
        "material_label": "金脉物料号/供应商料号",
    },
}

DEFAULT_LAYOUT = {
    "brand_lines": ("", ""),
    "footer_note": "",
    "doc_prefix_default": "WKT",
    "material_label": "物料号/供应商料号",
}


def layout_for(customer: str) -> dict:
    base = dict(DEFAULT_LAYOUT)
    base.update(CUSTOMER_LAYOUT.get((customer or "").strip(), {}))
    return base


def _set(ws, row: int, col: int, value: str, *, bold: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = LEFT if col <= 2 else CENTER
    cell.border = BORDER
    if bold:
        cell.font = Font(bold=True, size=14 if row <= 3 else 11)


def _merge_set(ws, row: int, col_start: int, col_end: int, value: str, *, bold: bool = False) -> None:
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    _set(ws, row, col_start, value, bold=bold)


def build_bilingual_delivery_workbook(customer: str) -> "Workbook":
    if Workbook is None:
        raise ValueError("openpyxl 未安装，无法生成专用送货单模板")

    cfg = layout_for(customer)
    brand1, brand2 = cfg["brand_lines"]
    material_label = cfg["material_label"]
    footer_note = cfg["footer_note"]

    wb = Workbook()
    ws = wb.active
    ws.title = "送货单"

    widths = [6, 16, 14, 20, 10, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 抬头
    ws.merge_cells("A1:D3")
    c = ws["A1"]
    c.value = f"{brand1}\n{brand2}".strip()
    c.alignment = LEFT
    c.font = Font(bold=True, size=11, color="0047AB")
    c.border = BORDER

    ws.merge_cells("E1:G3")
    t = ws["E1"]
    t.value = "送货单(Delivery Note)"
    t.alignment = CENTER
    t.font = Font(bold=True, size=16)
    t.border = BORDER

    # 单号 / 日期
    _set(ws, 4, 1, "Delivery No. 送货单号：")
    _merge_set(ws, 4, 2, 4, "{{送货单号}}")
    _set(ws, 4, 5, "Order Date 制单日期：")
    _merge_set(ws, 4, 6, 7, "{{制单日期}}")

    _set(ws, 5, 1, "W/H Location 仓库库位：")
    _merge_set(ws, 5, 2, 4, "{{仓库库位}}")
    _set(ws, 5, 5, "Delivery Date 发货日期：")
    _merge_set(ws, 5, 6, 7, "{{发货日期}}")

    # 供应商 / 客户
    _merge_set(ws, 6, 1, 4, "From: Supplier 供应商", bold=True)
    _merge_set(ws, 6, 5, 7, "To: Customer 客户", bold=True)

    supplier_rows = [
        ("Company 公司：", "{{供应商名称}}", "Company 公司：", "{{收货公司}}"),
        ("Address 地址：", "{{供应商地址}}", "Address 地址：", "{{收货地址}}"),
        ("Contacts 联系人：", "{{供应商联系人}}", "Contacts 联系人：", "{{收货联系人}}"),
        ("Telephone 电话：", "{{供应商电话}}", "Telephone 电话：", "{{收货电话}}"),
    ]
    for i, (sl, sv, cl, cv) in enumerate(supplier_rows, start=7):
        _set(ws, i, 1, sl)
        _merge_set(ws, i, 2, 4, sv)
        _set(ws, i, 5, cl)
        _merge_set(ws, i, 6, 7, cv)

    # 明细表头
    headers = [
        "Serial Number\n序号",
        material_label + "\nMaterial Numbers",
        "Supplier P/N\n供应商货号",
        "Description\n产品描述",
        "QTY(PCS)\n数量",
        "PO No.\nPO号",
        "Comments\n备注",
    ]
    ws.merge_cells("A11:A12")
    _set(ws, 11, 1, headers[0])
    for col, h in enumerate(headers[1:], start=2):
        ws.merge_cells(start_row=11, start_column=col, end_row=12, end_column=col)
        _set(ws, 11, col, h)

    data_cols = ["{{序号}}", "{{客户料号}}", "{{供应商货号}}", "{{产品描述}}", "{{本次出货}}", "{{订单号}}", "{{备注}}"]
    for col, val in enumerate(data_cols, start=1):
        _set(ws, 13, col, val)

    ws.merge_cells("A14:E14")
    _set(ws, 14, 1, "合计 Total")
    _merge_set(ws, 14, 6, 7, "{{合计}}")

    _merge_set(ws, 15, 1, 7, "Remarks 附注：{{附注}}")

    sign_rows = [
        ("Consigner 发货人：", "{{发货人}}", "Carrier 承运人：", "{{承运人}}"),
        ("Date 日期：", "{{发货日期}}", "Date 日期：", "{{承运日期}}"),
        ("Inspection 检验：", "{{检验}}", "Consignee 收货人：", "{{收货人}}"),
        ("Date 日期：", "{{检验日期}}", "Date 日期：", "{{收货日期}}"),
        ("Financial 财务：", "{{财务}}", "", ""),
        ("Date 日期：", "{{财务日期}}", "", ""),
    ]
    start = 16
    for i, (ll, lv, rl, rv) in enumerate(sign_rows):
        row = start + i
        _set(ws, row, 1, ll)
        _merge_set(ws, row, 2, 4, lv)
        if rl:
            _set(ws, row, 5, rl)
            _merge_set(ws, row, 6, 7, rv)

    if footer_note:
        ws.merge_cells("E21:G21")
        c = ws["E21"]
        c.value = footer_note
        c.alignment = LEFT
        c.border = BORDER

    return wb


def save_bilingual_template(customer: str, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_bilingual_delivery_workbook(customer)
    wb.save(path)
    return path
