"""送货单：威可特统一版式（HTML 打印 + Excel 下载）。"""
from __future__ import annotations

import io
import re
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from test_impl.common.money import serialize_qty
from test_impl.order_management.order_entry.line_service import OrderLineService
from test_impl.order_management.order_entry.shipment_models import ShipmentEvent

from .template_store import BUILTIN_HTML, DeliveryTemplateStore, WKT_STANDARD
from .wkt_document import (
    build_document_from_event,
    build_draft_document,
    build_sample_document,
    delivery_doc_prefix,
    document_from_dict,
    document_to_dict,
    get_customer_delivery_info,
    load_company_config,
    load_customer_delivery_config,
    save_customer_delivery_info,
)
from .wkt_document import _gen_doc_no  # noqa: PLC2701
from .wkt_xlsx import build_xlsx_bytes

try:
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Border, Side
except ImportError:
    load_workbook = None  # type: ignore
    MergedCell = None  # type: ignore
    Border = None  # type: ignore
    Side = None  # type: ignore

_THIN_SIDE = Side(style="thin", color="000000") if Side is not None else None
_FULL_BORDER = (
    Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
    if Border is not None and _THIN_SIDE is not None
    else None
)

_PLACEHOLDER_RE = re.compile(r"\{\{\s*([^}]+?)\s*\}\}")


def _fmt_dt_local(dt: datetime) -> str:
    if dt.tzinfo is not None:
        local = dt.astimezone()
    else:
        local = dt.replace(tzinfo=timezone.utc).astimezone()
    return local.strftime("%Y-%m-%d %H:%M")


def _fmt_date(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    if "T" in s:
        return s.split("T", 1)[0]
    return s[:10] if len(s) >= 10 else s


def _fmt_date_dot(s: str) -> str:
    s = _fmt_date(s)
    if not s or len(s) < 10:
        return s
    y, m, d = s.split("-", 2)
    try:
        return f"{y}.{int(m)}.{int(d)}"
    except ValueError:
        return s


def _split_contact_phone(text: str) -> Tuple[str, str]:
    text = (text or "").strip()
    if not text:
        return "", ""
    m = re.search(r"(\d{7,})", text)
    if m:
        phone = m.group(1)
        contact = text[: m.start()].strip(" ,，、")
        return contact, phone
    return text, ""


_LINE_PLACEHOLDER_KEYS = frozenset(
    {
        "序号",
        "客户料号",
        "供应商货号",
        "产品描述",
        "品名规格",
        "本次出货",
        "出货数量",
        "订单号",
        "备注",
        "order_no",
        "product_spec",
        "ship_qty",
        "customer_part_no",
        "材质",
        "单位",
    }
)

_TOTAL_PLACEHOLDER_KEYS = frozenset({"合计"})


def _row_line_placeholder_score(ws, row_idx: int) -> int:
    score = 0
    for cell in ws[row_idx]:
        val = cell.value
        if not isinstance(val, str) or "{{" not in val:
            continue
        for match in _PLACEHOLDER_RE.finditer(val):
            if match.group(1).strip() in _LINE_PLACEHOLDER_KEYS:
                score += 1
    return score


def _find_total_row(ws) -> Optional[int]:
    for row_idx in range(1, ws.max_row + 1):
        for cell in ws[row_idx]:
            val = cell.value
            if not isinstance(val, str) or "{{" not in val:
                continue
            for match in _PLACEHOLDER_RE.finditer(val):
                if match.group(1).strip() in _TOTAL_PLACEHOLDER_KEYS:
                    return row_idx
    return None


def _unmerge_row(ws, row_idx: int) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= row_idx <= merged.max_row:
            ws.unmerge_cells(str(merged))


def _row_has_content(ws, row_idx: int) -> bool:
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row_idx, col).value
        if val is not None and str(val).strip():
            return True
    return False


def _capture_row_template(ws, row_idx: int) -> Tuple[Dict[int, Any], List[Tuple[int, int]], Dict[int, Any]]:
    cells: Dict[int, Any] = {}
    styles: Dict[int, Any] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col)
        if cell.value is not None:
            cells[col] = cell.value
        if cell.has_style:
            styles[col] = {
                "font": copy(cell.font),
                "border": copy(cell.border),
                "fill": copy(cell.fill),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
            }
    merges: List[Tuple[int, int]] = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row == merged.max_row == row_idx:
            merges.append((merged.min_col, merged.max_col))
    return cells, merges, styles


def _apply_row_template(
    ws,
    row_idx: int,
    cells: Dict[int, Any],
    merges: List[Tuple[int, int]],
    styles: Dict[int, Any],
    replace_fn,
    ctx: Dict[str, str],
) -> None:
    _unmerge_row(ws, row_idx)
    for col, val in cells.items():
        cell = ws.cell(row_idx, col)
        if isinstance(val, str) and "{{" in val:
            cell.value = replace_fn(val, ctx)
        else:
            cell.value = val
        style = styles.get(col)
        if style:
            cell.font = copy(style["font"])
            cell.border = copy(style["border"])
            cell.fill = copy(style["fill"])
            cell.alignment = copy(style["alignment"])
            cell.number_format = style["number_format"]
    for min_col, max_col in merges:
        if max_col > min_col:
            ws.merge_cells(
                start_row=row_idx,
                start_column=min_col,
                end_row=row_idx,
                end_column=max_col,
            )


def _cell_has_border(cell) -> bool:
    border = cell.border
    return any(side and side.style for side in (border.left, border.right, border.top, border.bottom))


def _table_col_bounds(ws, detail_row: int) -> Tuple[int, int]:
    min_col = ws.max_column
    max_col = 1
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(detail_row, col)
        if cell.value is not None or _cell_has_border(cell):
            min_col = min(min_col, col)
            max_col = max(max_col, col)
    if min_col > max_col:
        return 1, max(1, ws.max_column)
    return min_col, max_col


def _apply_full_grid_borders(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    if _FULL_BORDER is None or MergedCell is None:
        return
    for row_idx in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row_idx, col)
            if isinstance(cell, MergedCell):
                continue
            cell.border = copy(_FULL_BORDER)
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row < min_row or merged.max_row > max_row:
            continue
        if merged.min_col < min_col or merged.max_col > max_col:
            continue
        top_left = ws.cell(merged.min_row, merged.min_col)
        if isinstance(top_left, MergedCell):
            continue
        top_left.border = copy(_FULL_BORDER)


def _copy_excel_row(ws, src_row: int, dst_row: int) -> None:
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst.value = src.value
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = copy(src.number_format)
            dst.protection = copy(src.protection)
            dst.alignment = copy(src.alignment)


def _find_detail_row(ws, *, before_row: Optional[int] = None) -> Optional[int]:
    limit = before_row if before_row is not None else ws.max_row + 1
    best_row: Optional[int] = None
    best_score = 0
    for row_idx in range(1, min(limit, ws.max_row + 1)):
        score = _row_line_placeholder_score(ws, row_idx)
        if score > best_score:
            best_score = score
            best_row = row_idx
    return best_row if best_score > 0 else None


class DeliveryNoteService:
    def __init__(
        self,
        line_service: OrderLineService,
        template_store: Optional[DeliveryTemplateStore] = None,
    ) -> None:
        self._lines = line_service
        self._templates = template_store or DeliveryTemplateStore()

    def get_event(self, event_id: int) -> ShipmentEvent:
        return self._lines.get_shipment_event(event_id)

    def build_context(self, event: ShipmentEvent) -> Dict[str, str]:
        line = self._lines.get_line(event.line_id)
        ship_qty = serialize_qty(event.ship_qty)
        po = serialize_qty(event.po_qty)
        shipped_after = serialize_qty(event.shipped_qty_after)
        open_after = serialize_qty(event.open_qty_after)
        shipped_at = _fmt_dt_local(event.shipped_at)
        ctx = {
            "customer": event.customer or line.customer,
            "order_no": event.order_no or line.order_no,
            "product_spec": event.product_spec or line.product_spec,
            "customer_part_no": event.customer_part_no or line.customer_part_no,
            "material": line.material or "",
            "unit": line.unit or "",
            "po_qty": po,
            "ship_qty": ship_qty,
            "shipped_qty_after": shipped_after,
            "open_qty_after": open_after,
            "order_date": _fmt_date(event.order_date or line.order_date),
            "delivery_date": _fmt_date(line.delivery_date),
            "shipped_at": shipped_at,
            "payment_terms": line.payment_terms or "",
            "客户": event.customer or line.customer,
            "订单号": event.order_no or line.order_no,
            "品名规格": event.product_spec or line.product_spec,
            "客户料号": event.customer_part_no or line.customer_part_no,
            "材质": line.material or "",
            "单位": line.unit or "",
            "PO数量": po,
            "本次出货": ship_qty,
            "出货数量": ship_qty,
            "累计已出货": shipped_after,
            "未结数量": open_after,
            "接单日期": _fmt_date(event.order_date or line.order_date),
            "客户交期": _fmt_date(line.delivery_date),
            "出货日期": shipped_at.split(" ")[0] if shipped_at else "",
            "出货时间": shipped_at,
            "账期": line.payment_terms or "",
        }
        return self._enrich_custom_template_context(
            ctx,
            (event.customer or line.customer or "").strip(),
            event=event,
            event_id=event.id,
        )

    def _enrich_custom_template_context(
        self,
        ctx: Dict[str, str],
        customer: str,
        *,
        event: Optional[ShipmentEvent] = None,
        event_id: Optional[int] = None,
    ) -> Dict[str, str]:
        from test_impl.order_management.customer_profile.store import get_profile

        customer = (customer or ctx.get("客户") or "").strip()
        company = load_company_config()
        delivery = get_customer_delivery_info(customer)
        profile = get_profile(customer)

        recv_contact = (delivery.get("receiver_contact") or "").strip()
        recv_phone = (delivery.get("receiver_phone") or "").strip()
        if not recv_contact and not recv_phone:
            recv_contact = (profile.get("contact") or "").strip()
            recv_phone = (profile.get("phone") or "").strip()
        elif not recv_phone:
            recv_contact, recv_phone = _split_contact_phone(delivery.get("receiver_contact", ""))

        prefix = delivery_doc_prefix(customer)

        ship_dt = event.shipped_at if event else datetime.now(timezone.utc)
        monthly_seq = 1
        if event is not None:
            eid = event_id if event_id is not None else getattr(event, "id", None)
            if eid:
                try:
                    monthly_seq = self._lines._store.monthly_sequence_for_shipment(eid, event.shipped_at)
                except Exception:
                    monthly_seq = 1

        order_date = ctx.get("接单日期") or ctx.get("order_date") or ""
        ship_date = ctx.get("出货日期") or order_date
        ship_dot = _fmt_date_dot(ship_date)
        recv_company = (delivery.get("receiver_company") or "").strip() or customer
        recv_address = (delivery.get("receiver_address") or "").strip()

        ctx.update(
            {
                "送货单号": _gen_doc_no(prefix, ship_dt, monthly_seq),
                "制单日期": _fmt_date_dot(order_date),
                "发货日期": ship_dot,
                "仓库库位": "",
                "供应商名称": company.get("supplier_name", ""),
                "供应商地址": company.get("supplier_address", ""),
                "供应商联系人": company.get("supplier_contact", ""),
                "供应商电话": company.get("supplier_phone", ""),
                "收货公司": recv_company,
                "收货地址": recv_address,
                "送货地点": recv_address,
                "订单下发抬头": recv_company,
                "收货联系人": recv_contact,
                "收货电话": recv_phone,
                "序号": "1",
                "供应商货号": ctx.get("材质") or ctx.get("客户料号") or "",
                "产品描述": ctx.get("品名规格") or "",
                "合计": ctx.get("本次出货") or "",
                "附注": "",
                "发货人": "",
                "检验": "",
                "检验日期": ship_dot,
                "财务": "",
                "财务日期": "",
                "承运人": "",
                "承运日期": "",
                "收货人": "",
                "收货日期": "",
                "备注": "",
            }
        )
        return ctx

    def build_batch_fill_contexts(
        self, event_ids: List[int]
    ) -> Tuple[Dict[str, str], List[Dict[str, str]]]:
        """合并出货：抬头共用一份 context，明细按出货记录各占一行。"""
        from decimal import Decimal

        from test_impl.common.money import round_qty, serialize_qty, to_decimal

        ids = [int(x) for x in event_ids if int(x) > 0]
        if not ids:
            raise ValueError("缺少出货记录")
        events = [self.get_event(eid) for eid in ids]
        if len(events) == 1:
            ctx = self.build_context(events[0])
            return ctx, [ctx]

        line_ctxs: List[Dict[str, str]] = []
        total = Decimal("0")
        for idx, ev in enumerate(events):
            line_ctx = self.build_context(ev)
            line_ctx["序号"] = str(idx + 1)
            line_ctx.pop("合计", None)
            line_ctxs.append(line_ctx)
            total += to_decimal(ev.ship_qty, field="本次出货")

        header = self.build_context(events[0])
        total_s = serialize_qty(round_qty(total))
        header["合计"] = total_s
        return header, line_ctxs

    def build_batch_context(self, event_ids: List[int]) -> Dict[str, str]:
        header, _ = self.build_batch_fill_contexts(event_ids)
        return header

    def build_wkt_document(self, event_id: int):
        snap = self._get_saved_delivery_note(event_id)
        if snap:
            return document_from_dict(snap)
        event = self.get_event(event_id)
        line = self._lines.get_line(event.line_id)
        monthly_seq = self._lines._store.monthly_sequence_for_shipment(event_id, event.shipped_at)
        return build_document_from_event(event, line, monthly_seq=monthly_seq)

    def build_wkt_document_dict(self, event_id: int) -> dict:
        return document_to_dict(self.build_wkt_document(event_id))

    def _get_saved_delivery_note(self, event_id: int) -> Optional[dict]:
        import json

        try:
            raw = self._lines._store.get_shipment_delivery_note_json(event_id)
        except ValueError:
            return None
        if not raw or not str(raw).strip():
            return None
        try:
            data = json.loads(raw)
            return data if isinstance(data, dict) else None
        except json.JSONDecodeError:
            return None

    def resolve_delivery_doc_no(
        self,
        event_id: int,
        delivery_note_json: str = "",
    ) -> str:
        """出货明细/对账用：优先读已保存 doc_no，否则按 WKT 规则推算。"""
        import json

        try:
            event = self.get_event(event_id)
        except ValueError:
            return ""

        customer = (event.customer or "").strip()
        expected_prefix = delivery_doc_prefix(customer)

        raw = (delivery_note_json or "").strip()
        if not raw:
            snap = self._get_saved_delivery_note(event_id)
        else:
            try:
                snap = json.loads(raw)
                if not isinstance(snap, dict):
                    snap = None
            except json.JSONDecodeError:
                snap = None
        if snap:
            doc_no = str(snap.get("doc_no") or "").strip()
            if doc_no.startswith(expected_prefix):
                return doc_no

        try:
            ui = self.ship_ui_mode(customer)
            if ui.get("mode") == "custom_excel":
                ctx = self.build_context(event)
                return str(ctx.get("送货单号") or "").strip()
            line = self._lines.get_line(event.line_id)
            monthly_seq = self._lines._store.monthly_sequence_for_shipment(event_id, event.shipped_at)
            doc = build_document_from_event(event, line, monthly_seq=monthly_seq)
            return str(doc.doc_no or "").strip()
        except ValueError:
            return ""

    def build_ship_draft(self, line_id: int, ship_qty: str) -> dict:
        line = self._lines.get_line(line_id)
        from test_impl.common.money import to_decimal

        qty = to_decimal(ship_qty, field="本次出货")
        doc = build_draft_document(line, qty)
        d = document_to_dict(doc)
        d["ship_qty"] = serialize_qty(qty)
        d["open_qty"] = serialize_qty(line.open_qty())
        d["customer"] = line.customer
        d["product_spec"] = line.product_spec
        return d

    def build_batch_ship_draft(self, items: List[dict]) -> dict:
        from test_impl.common.money import to_decimal
        from test_impl.order_management.delivery_note.wkt_document import build_batch_draft_document

        pairs: List[tuple] = []
        for raw in items:
            if not isinstance(raw, dict):
                raise ValueError("出货项格式无效")
            line_id = int(raw.get("line_id") or 0)
            if line_id <= 0:
                raise ValueError("缺少有效的 line_id")
            line = self._lines.get_line(line_id)
            qty = to_decimal(
                raw.get("qty") if raw.get("qty") is not None else raw.get("ship_qty"),
                field="本次出货",
            )
            pairs.append((line, qty))
        doc = build_batch_draft_document(pairs)
        d = document_to_dict(doc)
        d["customer"] = pairs[0][0].customer
        d["batch_items"] = [
            {"line_id": ln.id, "ship_qty": serialize_qty(q), "order_no": ln.order_no or ""}
            for ln, q in pairs
        ]
        return d

    def template_info(self, customer: str) -> dict:
        customer = (customer or "").strip()
        meta = self._templates.template_status(customer)
        return {"customer": customer, **meta, "is_excel": meta["is_custom_excel"], "is_builtin": False}

    def _require_custom_template_path(self, customer: str) -> Path:
        path = self._templates.resolve_template_path(customer)
        if path:
            return path
        status = self._templates.template_status(customer)
        if status["is_custom_excel"] and status["template_missing"]:
            raise ValueError(
                f"客户「{customer}」使用专用送货单模板「{status['template_file']}」，"
                f"但文件不存在。请将模板放入 data/delivery_templates/files/ 后重试。"
            )
        raise ValueError(f"客户「{customer}」未配置专用送货单模板")

    def _replace_placeholders(self, text: str, ctx: Dict[str, str]) -> str:
        if not text or "{{" not in text:
            return text

        def repl(m: re.Match) -> str:
            key = m.group(1).strip()
            return ctx.get(key, m.group(0))

        return _PLACEHOLDER_RE.sub(repl, str(text))

    def _fill_row_placeholders(self, ws, row_idx: int, ctx: Dict[str, str]) -> None:
        for cell in ws[row_idx]:
            val = cell.value
            if isinstance(val, str) and "{{" in val:
                cell.value = self._replace_placeholders(val, ctx)

    def _fill_worksheet(
        self,
        ws,
        ctx: Dict[str, str],
        line_contexts: Optional[List[Dict[str, str]]] = None,
    ) -> None:
        line_ctxs = line_contexts or []
        total_row = _find_total_row(ws)
        detail_row = _find_detail_row(ws, before_row=total_row)
        if detail_row and len(line_ctxs) > 1:
            extra = len(line_ctxs) - 1
            total_cells: Dict[int, Any] = {}
            total_merges: List[Tuple[int, int]] = []
            total_styles: Dict[int, Any] = {}
            remarks_snap: Optional[Tuple[int, Dict[int, Any], List[Tuple[int, int]], Dict[int, Any]]] = None
            new_total_row: Optional[int] = total_row
            insert_at = total_row if total_row and total_row > detail_row else detail_row + 1

            if total_row and total_row > detail_row:
                total_cells, total_merges, total_styles = _capture_row_template(ws, total_row)
                remarks_row = total_row + 1
                if remarks_row <= ws.max_row and _row_has_content(ws, remarks_row):
                    r_cells, r_merges, r_styles = _capture_row_template(ws, remarks_row)
                    remarks_snap = (remarks_row, r_cells, r_merges, r_styles)
                _unmerge_row(ws, total_row)
            if extra > 0:
                ws.insert_rows(insert_at, extra)
                for r in range(insert_at, insert_at + extra):
                    _unmerge_row(ws, r)
            for offset in range(1, len(line_ctxs)):
                _copy_excel_row(ws, detail_row, detail_row + offset)
            if total_row and total_row > detail_row:
                new_total_row = total_row + extra

            detail_end = detail_row + len(line_ctxs) - 1
            skip_rows = set()
            if new_total_row:
                skip_rows.add(new_total_row)
            new_remarks_row: Optional[int] = None
            if remarks_snap:
                new_remarks_row = remarks_snap[0] + extra
                skip_rows.add(new_remarks_row)

            for row_idx in range(1, ws.max_row + 1):
                if row_idx in skip_rows:
                    continue
                if detail_row <= row_idx <= detail_end:
                    row_ctx = {**ctx, **line_ctxs[row_idx - detail_row]}
                    self._fill_row_placeholders(ws, row_idx, row_ctx)
                else:
                    self._fill_row_placeholders(ws, row_idx, ctx)

            if new_total_row and total_cells:
                _apply_row_template(
                    ws,
                    new_total_row,
                    total_cells,
                    total_merges,
                    total_styles,
                    self._replace_placeholders,
                    ctx,
                )
            if remarks_snap and new_remarks_row:
                _, r_cells, r_merges, r_styles = remarks_snap
                _apply_row_template(
                    ws,
                    new_remarks_row,
                    r_cells,
                    r_merges,
                    r_styles,
                    self._replace_placeholders,
                    ctx,
                )
                for merged in ws.merged_cells.ranges:
                    if merged.min_row == merged.max_row == new_remarks_row:
                        top_left = ws.cell(merged.min_row, merged.min_col)
                        if not isinstance(top_left, MergedCell) and _FULL_BORDER is not None:
                            top_left.border = copy(_FULL_BORDER)

            if new_total_row and new_total_row >= detail_row:
                min_col, max_col = _table_col_bounds(ws, detail_row)
                _apply_full_grid_borders(ws, detail_row, new_total_row, min_col, max_col)
            return
        for row_idx in range(1, ws.max_row + 1):
            self._fill_row_placeholders(ws, row_idx, ctx)

    def fill_excel_bytes(
        self,
        template_path: Path,
        ctx: Dict[str, str],
        line_contexts: Optional[List[Dict[str, str]]] = None,
    ) -> bytes:
        if load_workbook is None:
            raise ValueError("openpyxl 未安装，无法生成 Excel 送货单")
        wb = load_workbook(template_path)
        for ws in wb.worksheets:
            self._fill_worksheet(ws, ctx, line_contexts)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def raw_template_bytes(self, customer: str) -> Tuple[bytes, str]:
        path = self._require_custom_template_path(customer)
        return path.read_bytes(), path.name

    def ship_ui_mode(self, customer: str) -> dict:
        from urllib.parse import quote

        from test_impl.order_management.customer_profile.store import get_profile, is_delivery_enabled

        customer = (customer or "").strip()
        if not customer:
            raise ValueError("请选择客户")
        if not is_delivery_enabled(get_profile(customer)):
            return {
                "mode": "none",
                "customer": customer,
                "label": "不使用送货单",
            }
        meta = self._templates.template_status(customer)
        q = quote(customer)
        if meta["is_wkt_standard"]:
            return {
                "mode": "wkt_standard",
                "customer": customer,
                "label": "威可特统一模板",
            }
        return {
            "mode": "custom_excel",
            "customer": customer,
            "label": f"专用模板 · {meta['template_file'] or 'Excel'}",
            "template_file": meta.get("template_file", ""),
            "template_missing": meta.get("template_missing", False),
            "raw_download_url": f"/api/delivery-templates/raw?customer={q}",
            "open_local": True,
        }

    def open_custom_excel_local(
        self,
        event_id: int,
        batch_event_ids: Optional[List[int]] = None,
        *,
        regenerate: bool = True,
    ) -> dict:
        event = self.get_event(event_id)
        customer = (event.customer or "").strip()
        custom_path = self._templates.resolve_template_path(customer)
        if not custom_path:
            raise ValueError("该客户未配置专用 Excel 模板")
        from .custom_excel_attachment import (
            attachment_path,
            open_in_excel,
            prepare_attachment_from_bytes,
            register_watch,
        )

        ids = [int(x) for x in (batch_event_ids or [event_id]) if int(x) > 0]
        if event_id not in ids:
            ids.insert(0, event_id)
        batch = len(ids) > 1
        existing = self._lines._store.get_shipment_attachment(event_id)

        if not regenerate and existing:
            path = attachment_path(existing)
            if path.is_file():
                register_watch(ids, existing, path)
                open_in_excel(path)
                return {
                    "ok": True,
                    "event_id": event_id,
                    "event_ids": ids,
                    "attachment": existing,
                    "path": str(path),
                    "message": "已在 Excel 中打开已保存的送货单",
                    "auto_filled": False,
                }

        if batch:
            header_ctx, line_ctxs = self.build_batch_fill_contexts(ids)
            doc_no = str(header_ctx.get("送货单号") or "").strip()
            filled = self.fill_excel_bytes(custom_path, header_ctx, line_ctxs)
        else:
            header_ctx = self.build_context(event)
            doc_no = str(header_ctx.get("送货单号") or "").strip()
            filled = self.fill_excel_bytes(custom_path, header_ctx)
        if doc_no:
            import json

            marker = json.dumps({"doc_no": doc_no}, ensure_ascii=False)
            for eid in ids:
                self._lines._store.save_shipment_delivery_note(int(eid), marker)
        rel, path = prepare_attachment_from_bytes(ids, filled, batch=batch)
        self._lines._store.save_shipment_attachment_batch(ids, rel)
        open_in_excel(path)
        return {
            "ok": True,
            "event_id": event_id,
            "event_ids": ids,
            "attachment": rel,
            "path": str(path),
            "message": "已自动填入订单数据并在 Excel 中打开；核对后保存即可写入出货明细",
            "auto_filled": True,
        }

    def attachment_status(self, event_id: int) -> dict:
        rel = self._lines._store.get_shipment_attachment(event_id)
        has_file = False
        mtime = 0.0
        if rel:
            from .custom_excel_attachment import attachment_path

            ap = attachment_path(rel)
            if ap.is_file():
                has_file = True
                mtime = ap.stat().st_mtime
        raw = self._lines._store.get_shipment_delivery_note_json(event_id)
        saved_at = ""
        if raw.strip():
            try:
                import json

                meta = json.loads(raw)
                if isinstance(meta, dict):
                    saved_at = str(meta.get("custom_excel_saved_at") or "")
            except json.JSONDecodeError:
                pass
        return {
            "ok": True,
            "event_id": event_id,
            "attachment": rel,
            "has_attachment": bool(rel and has_file),
            "attachment_mtime": mtime,
            "saved_at": saved_at,
        }

    def persist_attachment_saved(self, event_ids: List[int], rel_path: str) -> None:
        import json
        from datetime import datetime, timezone

        self._lines._store.save_shipment_attachment_batch(event_ids, rel_path)
        saved_at = datetime.now(timezone.utc).isoformat()
        for eid in event_ids:
            if int(eid) <= 0:
                continue
            meta = self._get_saved_delivery_note(int(eid)) or {}
            meta["custom_excel_saved_at"] = saved_at
            self._lines._store.save_shipment_delivery_note(
                int(eid),
                json.dumps(meta, ensure_ascii=False),
            )

    def render_for_event(self, event_id: int) -> Tuple[str, Any]:
        event = self.get_event(event_id)
        line = self._lines.get_line(event.line_id)
        customer = (event.customer or line.customer or "").strip()
        custom_path = self._templates.resolve_template_path(customer)
        if custom_path:
            rel = self._lines._store.get_shipment_attachment(event_id)
            if rel:
                from .custom_excel_attachment import attachment_path

                ap = attachment_path(rel)
                if ap.is_file():
                    return "xlsx", (ap.read_bytes(), ap.name)
            ctx = self.build_context(event)
            data = self.fill_excel_bytes(custom_path, ctx)
            fname = f"送货单_{_safe_file_part(customer)}_{event_id}.xlsx"
            return "xlsx", (data, fname)
        doc = self.build_wkt_document(event_id)
        data = build_xlsx_bytes(doc)
        fname = f"送货单_{_safe_file_part(event.customer)}_{event_id}.xlsx"
        return "xlsx", (data, fname)

    def build_sample_context(self, customer: str) -> Dict[str, str]:
        """维护页预览用示例数据（非真实出货）。"""
        customer = (customer or "").strip()
        now = datetime.now(timezone.utc).astimezone()
        shipped_at = now.strftime("%Y-%m-%d %H:%M")
        ship_date = now.strftime("%Y-%m-%d")
        ctx = {
            "customer": customer,
            "order_no": "PO-预览-001",
            "product_spec": "（示例品名规格）",
            "customer_part_no": "（示例料号）",
            "material": "（示例材质）",
            "unit": "PCS",
            "po_qty": "1000",
            "ship_qty": "100",
            "shipped_qty_after": "100",
            "open_qty_after": "900",
            "order_date": ship_date,
            "delivery_date": ship_date,
            "shipped_at": shipped_at,
            "payment_terms": "（示例账期）",
            "客户": customer,
            "订单号": "PO-预览-001",
            "品名规格": "（示例品名规格）",
            "客户料号": "（示例料号）",
            "材质": "（示例材质）",
            "单位": "PCS",
            "PO数量": "1000",
            "本次出货": "100",
            "出货数量": "100",
            "累计已出货": "100",
            "未结数量": "900",
            "接单日期": ship_date,
            "客户交期": ship_date,
            "出货日期": ship_date,
            "出货时间": shipped_at,
            "账期": "（示例账期）",
        }
        return self._enrich_custom_template_context(ctx, customer)

    def preview_for_customer(self, customer: str) -> dict:
        customer = (customer or "").strip()
        if not customer:
            raise ValueError("请选择客户")
        from test_impl.order_management.customer_profile.store import get_profile, is_delivery_enabled

        if not is_delivery_enabled(get_profile(customer)):
            raise ValueError("该客户未启用送货单")
        from urllib.parse import quote

        q = quote(customer)
        info = get_customer_delivery_info(customer)
        meta = self._templates.template_status(customer)
        out = {
            "customer": customer,
            **meta,
            "is_excel": meta["is_custom_excel"],
            "is_builtin": False,
            "preview_download_url": f"/api/delivery-templates/preview-download?customer={q}",
            "raw_download_url": f"/api/delivery-templates/raw?customer={q}",
            "sample_context": self.build_sample_context(customer),
            "delivery_info": info,
        }
        if meta["is_wkt_standard"]:
            out["preview_html_url"] = f"/delivery-note/preview-sample?customer={q}"
            out["template_label"] = "威可特统一送货单"
        else:
            out["preview_html_url"] = ""
            label = meta["template_file"] or "专用 Excel 模板"
            if meta["template_missing"]:
                label += "（模板文件待上传）"
            out["template_label"] = f"专用模板 · {label}"
            out["auto_fill"] = True
            out["placeholder_fields"] = self.list_placeholder_fields()
        return out

    def list_placeholder_fields(self) -> List[str]:
        return [
            "客户",
            "订单号",
            "品名规格",
            "客户料号",
            "材质",
            "单位",
            "本次出货",
            "出货数量",
            "PO数量",
            "累计已出货",
            "未结数量",
            "出货日期",
            "出货时间",
            "接单日期",
            "客户交期",
            "账期",
            "送货单号",
            "制单日期",
            "发货日期",
            "收货公司",
            "收货地址",
            "送货地点",
            "订单下发抬头",
            "收货联系人",
            "收货电话",
            "供应商名称",
            "供应商地址",
            "供应商电话",
            "合计",
            "产品描述",
            "供应商货号",
        ]

    def render_sample_for_customer(self, customer: str) -> Tuple[str, Any]:
        customer = (customer or "").strip()
        custom_path = self._templates.resolve_template_path(customer)
        if custom_path:
            ctx = self.build_sample_context(customer)
            data = self.fill_excel_bytes(custom_path, ctx)
            fname = f"送货单预览_{_safe_file_part(customer)}.xlsx"
            return "xlsx", (data, fname)
        doc = build_sample_document(customer)
        data = build_xlsx_bytes(doc)
        fname = f"送货单预览_{_safe_file_part(customer)}.xlsx"
        return "xlsx", (data, fname)

    def render_sample_html_doc(self, customer: str) -> dict:
        return document_to_dict(build_sample_document(customer))

    def list_config(self) -> dict:
        customer_names: set[str] = set()
        try:
            for name in self._lines.list_master().get("customers") or []:
                if str(name).strip():
                    customer_names.add(str(name).strip())
        except Exception:
            pass
        delivery_cfg = load_company_config()
        all_delivery = load_customer_delivery_config()
        mapping = self._templates.load_mapping()
        customer_names.update(all_delivery.keys())
        customer_names.update(mapping.keys())
        from test_impl.order_management.customer_profile.store import (
            get_profile,
            is_delivery_enabled,
            load_all_profiles,
        )
        from test_impl.order_management.reconciliation.period import reconciliation_period_label

        profiles = load_all_profiles()
        customer_names.update(profiles.keys())
        rows = []
        for name in sorted(customer_names, key=lambda x: x.lower()):
            info = get_customer_delivery_info(name)
            profile = profiles.get(name) or get_profile(name)
            delivery_on = is_delivery_enabled(profile)
            contact = (info.get("receiver_contact") or profile.get("contact") or "").strip()
            phone = (info.get("receiver_phone") or profile.get("phone") or "").strip()
            if contact and not phone:
                from .wkt_document import split_receiver_contact

                contact, phone = split_receiver_contact(contact)
            address = (info.get("receiver_address") or profile.get("address") or "").strip()
            meta = self._templates.template_status(name)
            if not delivery_on:
                template_display = "不使用送货单"
            elif meta["is_wkt_standard"]:
                template_display = "威可特统一模板"
            elif meta["template_missing"]:
                template_display = f"专用 · {meta['template_file']}（待放入）"
            else:
                template_display = f"专用 · {meta['template_file']}"
            rows.append(
                {
                    "customer": name,
                    "template": meta["template"],
                    "template_display": template_display,
                    "template_file": meta.get("template_file", ""),
                    "is_builtin": False,
                    "is_excel": meta["is_custom_excel"],
                    "is_wkt_standard": meta["is_wkt_standard"],
                    "is_custom_excel": meta["is_custom_excel"],
                    "template_missing": meta["template_missing"],
                    "delivery_enabled": delivery_on,
                    "receiver_address": address,
                    "receiver_contact": contact,
                    "receiver_phone": phone,
                    "doc_no_prefix": info.get("doc_no_prefix", ""),
                    "address": address,
                    "email": (profile.get("email") or "").strip(),
                    "payment_terms": (profile.get("payment_terms") or "").strip(),
                    "reconciliation_period": (profile.get("reconciliation_period") or "").strip(),
                    "reconciliation_period_label": reconciliation_period_label(
                        profile.get("reconciliation_period")
                    ),
                }
            )
        return {
            "template": WKT_STANDARD,
            "supplier": delivery_cfg,
            "builtin": BUILTIN_HTML,
            "mapping": mapping,
            "special_customers": list(mapping.keys()),
            "template_files": self._templates.list_template_files(),
            "placeholder_help": list(_wkt_field_help()),
            "customer_rows": rows,
            "customer_profiles": {
                k: {
                    "address": v.get("address", ""),
                    "contact": v.get("contact", ""),
                    "phone": v.get("phone", ""),
                    "email": v.get("email", ""),
                    "payment_terms": v.get("payment_terms", ""),
                    "reconciliation_period": v.get("reconciliation_period", ""),
                    "reconciliation_period_label": reconciliation_period_label(
                        v.get("reconciliation_period")
                    ),
                    "delivery_enabled": is_delivery_enabled(v),
                }
                for k, v in profiles.items()
            },
        }

    def get_customer_delivery(self, customer: str) -> dict:
        return get_customer_delivery_info(customer)

    def save_customer_delivery(self, customer: str, info: dict) -> None:
        save_customer_delivery_info(customer, info)

    def upload_template_file(self, filename: str, data: bytes) -> str:
        return self._templates.save_upload(filename, data)

    def set_customer_template(self, customer: str, template: str) -> None:
        self._templates.set_customer_template(customer, template)

    def remove_customer_template(self, customer: str) -> None:
        self._templates.remove_customer_mapping(customer)


def _safe_file_part(s: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "_", (s or "").strip())[:40] or "customer"


def _wkt_field_help() -> tuple[str, ...]:
    return (
        "专用 Excel：在单元格写 {{占位符}}，出货时系统自动替换。",
        "常用：{{客户}} {{订单号}} {{品名规格}} {{客户料号}} {{本次出货}} {{出货日期}}",
        "抬头/收货：{{送货单号}} {{收货地址}} {{送货地点}} {{订单下发抬头}} {{收货联系人}} {{收货电话}} {{供应商名称}}",
        "合并出货时：明细行占位符会按料号自动扩展为多行；{{合计}} 为总数量。",
        "预览可下载「已填入示例数据」的 Excel 核对占位位置是否正确。",
    )
