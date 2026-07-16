"""按威可特统一版式生成 Excel 送货单（对齐 HTML 预览/打印版式）。"""
from __future__ import annotations

import io
from typing import TYPE_CHECKING

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore

if TYPE_CHECKING:
    from .wkt_document import WktDeliveryDocument

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_LABEL = Alignment(horizontal="left", vertical="center", wrap_text=False)

_COL_WIDTHS = [14, 14, 16, 14, 6, 8, 12, 8, 12]


def _set_col_widths(ws) -> None:
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _merge_write(ws, cell_range: str, value, *, align=None, bold=False, border=False):
    ws.merge_cells(cell_range)
    top_left = cell_range.split(":")[0]
    cell = ws[top_left]
    cell.value = value
    cell.alignment = align or _LEFT
    cell.font = Font(bold=bold)
    if border:
        cell.border = _BORDER


def _write_meta_row(ws, row: int, left_lbl: str, left_val: str, right_lbl: str, right_val: str) -> None:
    ws.cell(row=row, column=1, value=left_lbl).alignment = _LABEL
    _merge_write(ws, f"B{row}:E{row}", left_val or "", align=_LEFT)
    ws.cell(row=row, column=6, value=right_lbl).alignment = _LABEL
    _merge_write(ws, f"G{row}:I{row}", right_val or "", align=_LEFT)


def build_xlsx_bytes(doc: "WktDeliveryDocument") -> bytes:
    if Workbook is None:
        raise ValueError("openpyxl 未安装，无法生成 Excel 送货单")
    wb = Workbook()
    ws = wb.active
    ws.title = "送货单"
    _set_col_widths(ws)

    _merge_write(ws, "A1:I1", doc.title_company, align=_CENTER, bold=True)
    ws["A1"].font = Font(size=16, bold=True)
    _merge_write(ws, "A2:I2", "送货单", align=_CENTER, bold=True)
    ws["A2"].font = Font(size=14, bold=True)

    _write_meta_row(ws, 4, "送货单号：", doc.doc_no, "日期：", doc.ship_date_cn)
    _write_meta_row(ws, 5, "收货公司：", doc.receiver_company, "供应商：", doc.supplier_name)
    _write_meta_row(ws, 6, "收货地点：", doc.receiver_address, "供应商地址：", doc.supplier_address)
    _write_meta_row(ws, 7, "联系人及电话：", doc.receiver_contact, "供应商电话：", doc.supplier_phone)
    ws.row_dimensions[6].height = 36

    headers = [
        "订单号",
        "客户物料编码",
        "物料名称",
        "规格型号",
        "单位",
        "数量",
        "生产批号",
        "箱数",
        "备注",
    ]
    start_row = 9
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = _CENTER
        cell.border = _BORDER

    row = start_row + 1
    for ln in doc.lines:
        vals = [
            ln.order_no or "/",
            ln.customer_part_no,
            ln.product_name,
            ln.spec,
            ln.unit,
            ln.qty,
            ln.batch_no,
            ln.box_count,
            ln.remark,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = _BORDER
            cell.alignment = _CENTER if col in (5, 6, 7, 8) else _LEFT
        row += 1

    _merge_write(ws, f"A{row}:I{row}", "以下空白", align=_LEFT, border=True)
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = _BORDER
    row += 1

    _merge_write(
        ws,
        f"A{row}:E{row}",
        "合计",
        align=Alignment(horizontal="left", vertical="center"),
        bold=True,
    )
    qty_cell = ws.cell(row=row, column=6, value=doc.total_qty)
    qty_cell.font = Font(bold=True)
    qty_cell.alignment = _CENTER
    for col in range(1, 10):
        ws.cell(row=row, column=col).border = _BORDER

    sign_row = row + 2
    _merge_write(ws, f"A{sign_row}:C{sign_row}", "送货人：" + (doc.deliverer or ""), align=_LEFT)
    _merge_write(ws, f"D{sign_row}:F{sign_row}", "仓管：" + (doc.warehouse_manager or ""), align=_LEFT)
    _merge_write(ws, f"G{sign_row}:I{sign_row}", "收货：" + (doc.receiver_sign or ""), align=_LEFT)

    _merge_write(ws, f"A{sign_row + 2}:I{sign_row + 2}", doc.footer_note or "", align=_LEFT)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
