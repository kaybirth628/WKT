"""Excel/CSV 订单行批量导入与校验。"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from test_impl.common.money import OPEN_QTY_EPS, round_amount, round_qty, to_decimal
from test_impl.order_management.order_entry.line_models import normalize_line_fields

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

# 表头别名 → 系统字段（小写、去空格后匹配）
HEADER_ALIASES: Dict[str, str] = {
    "客户": "customer",
    "客户名称": "customer",
    "客户名": "customer",
    "接单日期": "order_date",
    "订单日期": "order_date",
    "客户交期": "delivery_date",
    "交期": "delivery_date",
    "订单号": "order_no",
    "po号": "order_no",
    "采购订单号": "order_no",
    "品名规格": "product_spec",
    "品名/规格": "product_spec",
    "品名": "product_spec",
    "客户料号": "customer_part_no",
    "料号": "customer_part_no",
    "单重": "unit_weight_g",
    "单重g": "unit_weight_g",
    "单重（不含损耗）g": "unit_weight_g",
    "单重(不含损耗)g": "unit_weight_g",
    "材质": "material",
    "po数量": "po_qty",
    "数量": "po_qty",
    "订购数量": "po_qty",
    "已出货": "shipped_qty",
    "已出货数量": "shipped_qty",
    "出货数量": "shipped_qty",
    "出货": "shipped_qty",
    "已出数量": "shipped_qty",
    "未结数量": "open_qty",
    "未出货数量": "open_qty",
    "未出货": "open_qty",
    "未结": "open_qty",
    "未结及出货": "open_qty",
    "未结数量及出货": "open_qty",
    "订单未结数量": "open_qty",
    "剩余数量": "open_qty",
    "结余数量": "open_qty",
    "单位": "unit",
    "税率": "tax_rate",
    "税率%": "tax_rate",
    "人民币单价（含税）": "rmb_tax_incl_price",
    "人民币单价(含税)": "rmb_tax_incl_price",
    "人民币含税单价": "rmb_tax_incl_price",
    "人民币单价": "rmb_tax_incl_price",
    "含税单价": "rmb_tax_incl_price",
    "单价": "rmb_tax_incl_price",
    "人民币金额": "amount",
    "人民币合计": "amount",
    "账期": "payment_terms",
    "付款条件": "payment_terms",
    "结算方式": "payment_terms",
    "账期说明": "payment_terms",
    "付款账期": "payment_terms",
    "金额": "amount",
    "含税金额": "amount",
    "行金额": "amount",
    "订单金额": "amount",
    "总金额": "amount",
    "价税合计": "amount",
    "含税总价": "amount",
    "销售额": "amount",
}

AMOUNT_EPS = Decimal("0.01")

# 与业务台账表头一致（15 列）；「金额」为可选扩展列，不在模板中
TEMPLATE_HEADERS = [
    "客户",
    "接单日期",
    "客户交期",
    "订单号",
    "品名规格",
    "客户料号",
    "单重（不含损耗）g",
    "材质",
    "PO数量",
    "已出货",
    "未结数量",
    "单位",
    "税率",
    "人民币单价（含税）",
    "账期",
]


def _norm_header(text: str) -> str:
    s = str(text or "").strip().lower()
    s = re.sub(r"\s+", "", s)
    return s


def _resolve_header_field(key: str) -> Optional[str]:
    if not key:
        return None
    if key in HEADER_ALIASES:
        return HEADER_ALIASES[key]
    if key.startswith("人民币单价") or key.startswith("人民币含税"):
        return "rmb_tax_incl_price"
    if key.startswith("人民币") and any(x in key for x in ("金额", "合计", "总价", "价税")):
        return "amount"
    return None


def _humanize_error_message(msg: str) -> str:
    """把 Decimal 底层异常转成业务可读说明。"""
    low = (msg or "").lower()
    if "conversionsyntax" in low or "class 'decimal" in low or "invalidoperation" in low:
        return (
            "数字列含有无法识别的内容，请检查："
            "PO数量、已出货、未结数量、税率、人民币单价（含税）是否为数字"
        )
    return msg


def _friendly_exception(exc: BaseException) -> str:
    if isinstance(exc, ValueError):
        return _humanize_error_message(str(exc))
    if isinstance(exc, InvalidOperation):
        return _humanize_error_message(str(exc))
    return _humanize_error_message(str(exc))


def map_headers(raw_headers: List[str]) -> Tuple[Dict[int, str], List[str]]:
    """返回 {列索引: 字段名} 与未识别的表头列表。"""
    mapping: Dict[int, str] = {}
    unknown: List[str] = []
    for idx, h in enumerate(raw_headers):
        key = _norm_header(h)
        if not key:
            continue
        field_name = _resolve_header_field(key)
        if field_name:
            mapping[idx] = field_name
        else:
            unknown.append(str(h).strip())
    return mapping, unknown


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _row_is_empty(cells: Dict[str, str]) -> bool:
    keys = ("customer", "order_no", "product_spec", "po_qty")
    return not any(cells.get(k) for k in keys)


@dataclass
class ImportIssue:
    field: str
    level: str  # error | warn
    message: str

    def __post_init__(self) -> None:
        self.message = _humanize_error_message(self.message)

    def to_dict(self) -> dict:
        return {
            "field": self.field,
            "level": self.level,
            "message": self.message,
        }


NUMERIC_FIELD_LABELS: List[Tuple[str, str]] = [
    ("po_qty", "PO数量"),
    ("shipped_qty", "已出货"),
    ("rmb_tax_incl_price", "人民币单价（含税）"),
    ("tax_rate", "税率"),
]
# 单重（不含损耗）g：允许文字备注如「外购件」，不在此做数字强校验


def _check_raw_numeric_fields(raw: dict) -> List[ImportIssue]:
    """逐列检查数字字段，返回带列名的可读错误。"""
    from test_impl.order_management.order_entry.line_models import _normalize_tax_rate

    issues: List[ImportIssue] = []
    for key, label in NUMERIC_FIELD_LABELS:
        val = raw.get(key)
        s = str(val or "").strip()
        if not s:
            continue
        try:
            if key == "tax_rate":
                _normalize_tax_rate(val)
            else:
                to_decimal(val, field=label)
        except (ValueError, InvalidOperation) as exc:
            issues.append(ImportIssue(key, "error", _friendly_exception(exc)))
    return issues


@dataclass
class ImportRowResult:
    row_no: int
    data: dict
    calc_amount: str = ""
    excel_amount: str = ""
    calc_open_qty: str = ""
    excel_open_qty: str = ""
    issues: List[ImportIssue] = field(default_factory=list)

    @property
    def importable(self) -> bool:
        return not any(i.level == "error" for i in self.issues)

    @property
    def review_status(self) -> str:
        """passed=可直接导入；pending=待确认(有警告)；blocked=阻断。"""
        if not self.importable:
            return "blocked"
        if any(i.level == "warn" for i in self.issues):
            return "pending"
        return "passed"

    def to_dict(self) -> dict:
        return {
            "row_no": self.row_no,
            "data": self.data,
            "calc_amount": self.calc_amount,
            "excel_amount": self.excel_amount,
            "calc_open_qty": self.calc_open_qty,
            "excel_open_qty": self.excel_open_qty,
            "issues": [i.to_dict() for i in self.issues],
            "importable": self.importable,
            "review_status": self.review_status,
        }


def _looks_numeric_tax(value: str) -> bool:
    s = str(value or "").strip().replace("%", "").replace(",", "")
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def _looks_like_payment_terms(value: str) -> bool:
    s = str(value or "").strip()
    if not s or _looks_numeric_tax(s):
        return False
    hints = ("月结", "开票", "账期", "付款", "结算", "货到付款", "次月", "当月", "天")
    return any(h in s for h in hints)


def _prepare_import_raw(raw: dict) -> dict:
    """Excel 账期按原文保存；若账期说明误入「税率」列则自动纠正。"""
    out = dict(raw)
    pt = str(out.get("payment_terms") or "").strip()
    tr = str(out.get("tax_rate") or "").strip()
    if tr and _looks_like_payment_terms(tr):
        if not pt:
            out["payment_terms"] = tr
            out["tax_rate"] = ""
        elif not _looks_numeric_tax(tr):
            out["payment_terms"] = pt if len(pt) >= len(tr) else tr
            out["tax_rate"] = ""
    if "payment_terms" in out:
        out["payment_terms"] = str(out["payment_terms"]).strip()[:200]
    return out


def normalize_line_fields_for_import(data: dict) -> tuple[dict, List[ImportIssue]]:
    """导入专用：账期原文保留；税率无法解析时按 0 并提示。"""
    prep = _prepare_import_raw(data)
    extra: List[ImportIssue] = []
    try:
        return normalize_line_fields(prep), extra
    except (ValueError, InvalidOperation) as exc:
        msg = _friendly_exception(exc)
        if "税率" in msg:
            prep["tax_rate"] = ""
            extra.append(
                ImportIssue(
                    "tax_rate",
                    "warn",
                    f"税率无法按数字解析，已按 0 处理（原文：{prep.get('tax_rate') or data.get('tax_rate')!r}）",
                )
            )
            return normalize_line_fields(prep), extra
        raise ValueError(msg) from exc


def validate_import_row(raw: dict, *, row_no: int) -> ImportRowResult:
    """校验单行：业务规则 + 未结数量 + 金额。"""
    excel_open = raw.pop("_excel_open_qty", None)
    excel_amount = raw.pop("_excel_amount", None)
    raw = _prepare_import_raw(raw)

    num_issues = _check_raw_numeric_fields(raw)
    if any(i.level == "error" for i in num_issues):
        result = ImportRowResult(
            row_no=row_no,
            data={k: str(v) for k, v in raw.items()},
        )
        result.issues.extend(num_issues)
        return result

    try:
        fields, prep_issues = normalize_line_fields_for_import(raw)
    except (ValueError, InvalidOperation) as exc:
        result = ImportRowResult(row_no=row_no, data=raw)
        result.issues.append(ImportIssue("row", "error", _friendly_exception(exc)))
        return result

    po = fields["po_qty"]
    shipped = fields["shipped_qty"]
    price = fields["rmb_tax_incl_price"]
    calc_open = round_qty(po - shipped)
    calc_amt = round_amount(po * price)

    result = ImportRowResult(
        row_no=row_no,
        data={
            k: (str(v) if isinstance(v, Decimal) else v) for k, v in fields.items()
        },
        calc_amount=str(calc_amt),
        calc_open_qty=str(calc_open),
    )
    result.issues.extend(prep_issues)

    if excel_open not in (None, ""):
        result.excel_open_qty = str(excel_open)
        try:
            diff = abs(round_qty(to_decimal(excel_open, field="未结数量")) - calc_open)
        except (ValueError, InvalidOperation) as exc:
            result.issues.append(ImportIssue("open_qty", "error", _friendly_exception(exc)))
        else:
            if diff > OPEN_QTY_EPS:
                result.issues.append(
                    ImportIssue(
                        "open_qty",
                        "error",
                        f"未结数量不一致：Excel={excel_open}，应为 PO−已出货={calc_open}",
                    )
                )
    elif po > 0 and shipped > po:
        pass  # caught below

    if excel_amount not in (None, ""):
        result.excel_amount = str(excel_amount)
        try:
            excel_amt = round_amount(to_decimal(excel_amount, field="金额"))
            diff_amt = abs(excel_amt - calc_amt)
        except (ValueError, InvalidOperation) as exc:
            result.issues.append(ImportIssue("amount", "error", _friendly_exception(exc)))
        else:
            if diff_amt > AMOUNT_EPS:
                result.issues.append(
                    ImportIssue(
                        "amount",
                        "error",
                        f"金额不一致：Excel={excel_amt}，计算值 PO×单价={calc_amt}",
                    )
                )

    # 必填与 OrderLine 规则
    try:
        from test_impl.order_management.order_entry.line_models import OrderLine

        probe = OrderLine(id=0, **fields)
        probe.validate()
    except ValueError as exc:
        result.issues.append(ImportIssue("row", "error", _friendly_exception(exc)))

    if not fields.get("customer"):
        result.issues.append(ImportIssue("customer", "error", "客户不能为空"))
    if not fields.get("order_no"):
        result.issues.append(ImportIssue("order_no", "error", "订单号不能为空"))
    if not fields.get("product_spec"):
        result.issues.append(ImportIssue("product_spec", "error", "品名规格不能为空"))

    return result


def _parse_sheet_rows(matrix: List[List[Any]], *, start_row: int = 1) -> Tuple[List[Tuple[int, dict]], List[str]]:
    if not matrix:
        raise ValueError("表格为空")
    headers = [_cell_str(c) for c in matrix[0]]
    col_map, unknown = map_headers(headers)
    if not col_map:
        raise ValueError(
            "未识别到有效表头，请使用导入模板或确保首行包含「客户」「订单号」「品名规格」等列"
        )

    parsed: List[Tuple[int, dict]] = []
    for ridx, row in enumerate(matrix[start_row:], start=start_row + 1):
        cells: Dict[str, str] = {}
        for cidx, field_name in col_map.items():
            val = row[cidx] if cidx < len(row) else None
            cells[field_name] = _cell_str(val)
        if _row_is_empty(cells):
            continue
        raw = {k: v for k, v in cells.items() if k not in ("open_qty", "amount")}
        if cells.get("open_qty"):
            raw["_excel_open_qty"] = cells["open_qty"]
        if cells.get("amount"):
            raw["_excel_amount"] = cells["amount"]
        parsed.append((ridx, raw))

    return parsed, unknown


def _results_from_parsed(parsed: List[Tuple[int, dict]], unknown: List[str]) -> List[ImportRowResult]:
    return [validate_import_row(dict(raw), row_no=row_no) for row_no, raw in parsed]


def parse_excel_bytes(data: bytes, filename: str = "") -> Tuple[List[ImportRowResult], List[str]]:
    if load_workbook is None:
        raise ValueError("服务器未安装 openpyxl，请执行 pip install openpyxl")
    ext = Path(filename or "").suffix.lower()
    if ext == ".csv":
        return parse_csv_bytes(data)

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    matrix = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    parsed, unknown = _parse_sheet_rows(matrix)
    return _results_from_parsed(parsed, unknown), unknown


def parse_csv_bytes(data: bytes, encoding: str = "utf-8-sig") -> Tuple[List[ImportRowResult], List[str]]:
    text = data.decode(encoding, errors="replace")
    reader = csv.reader(io.StringIO(text))
    matrix = list(reader)
    parsed, unknown = _parse_sheet_rows(matrix)
    return _results_from_parsed(parsed, unknown), unknown


def _row_brief(r: ImportRowResult) -> dict:
    d = r.data or {}
    err_msgs = [_humanize_error_message(i.message) for i in r.issues if i.level == "error"]
    warn_msgs = [_humanize_error_message(i.message) for i in r.issues if i.level == "warn"]
    return {
        "row_no": r.row_no,
        "review_status": r.review_status,
        "customer": d.get("customer", ""),
        "order_no": d.get("order_no", ""),
        "product_spec": d.get("product_spec", ""),
        "po_qty": d.get("po_qty", ""),
        "shipped_qty": d.get("shipped_qty", ""),
        "payment_terms": d.get("payment_terms", ""),
        "excel_open_qty": r.excel_open_qty,
        "calc_open_qty": r.calc_open_qty,
        "errors": err_msgs,
        "warnings": warn_msgs,
    }


def build_blocked_report(blocked_list: List[dict]) -> str:
    """生成可复制的阻断原因全文。"""
    if not blocked_list:
        return ""
    parts: List[str] = [
        f"Excel 导入阻断报告（共 {len(blocked_list)} 行）",
        "=" * 40,
        "",
    ]
    for b in blocked_list:
        parts.append(f"【Excel 第 {b.get('row_no')} 行】")
        parts.append(f"  客户：{b.get('customer', '')}")
        parts.append(f"  订单号：{b.get('order_no', '')}")
        parts.append(f"  品名规格：{b.get('product_spec', '')}")
        parts.append(f"  PO数量：{b.get('po_qty', '')}  已出货：{b.get('shipped_qty', '')}")
        if b.get("excel_open_qty"):
            parts.append(
                f"  未结数量(Excel)：{b.get('excel_open_qty')}  系统计算：{b.get('calc_open_qty', '')}"
            )
        for e in b.get("errors") or []:
            parts.append(f"  ✗ {e}")
        parts.append("")
    return "\n".join(parts)


def summarize_results(
    rows: List[ImportRowResult], *, unknown_headers: Optional[List[str]] = None
) -> dict:
    importable = sum(1 for r in rows if r.importable)
    errors = sum(1 for r in rows if any(i.level == "error" for i in r.issues))
    warnings = sum(
        1 for r in rows if r.importable and any(i.level == "warn" for i in r.issues)
    )
    passed_rows = [r for r in rows if r.review_status == "passed"]
    pending_rows = [r for r in rows if r.review_status == "pending"]
    blocked_rows = [r for r in rows if r.review_status == "blocked"]
    out = {
        "total": len(rows),
        "importable": importable,
        "passed": len(passed_rows),
        "pending": len(pending_rows),
        "blocked": len(blocked_rows),
        "error_rows": errors,
        "warn_rows": warnings,
        "passed_list": [_row_brief(r) for r in passed_rows],
        "pending_list": [_row_brief(r) for r in pending_rows],
        "blocked_list": [_row_brief(r) for r in blocked_rows],
        "rows": [r.to_dict() for r in rows],
        "blocked_report": build_blocked_report([_row_brief(r) for r in blocked_rows]),
    }
    if unknown_headers:
        out["header_warnings"] = [
            f"以下列名未纳入导入（可删除或改与模板一致）：{', '.join(unknown_headers[:12])}"
        ]
    return out


def build_template_bytes() -> bytes:
    if Workbook is None:
        raise ValueError("openpyxl 未安装")
    wb = Workbook()
    ws = wb.active
    ws.title = "订单导入"
    ws.append(TEMPLATE_HEADERS)
    ws.append(
        [
            "华东精密机械",
            "2026-05-12",
            "2026-06-12",
            "PO-DEMO-001",
            "散热片/8513 A",
            "B603000053",
            45,
            "ADC12",
            1000,
            300,
            700,
            "PCS",
            "13%",
            1.8,
            "月结60天",
        ]
    )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
