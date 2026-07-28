"""解析威可特「新产品 BOM 表」表单式 Excel（每 sheet 一个料号）。"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from typing import Any, Callable, Dict, List, Optional, Tuple

from test_impl.order_management.customer_name import (
    _FILENAME_CUSTOMER_ALIASES,
    extract_customer_hint_from_filename,
    normalize_customer_name,
    resolve_customer_from_hint,
    dedupe_customer_names,
)
from test_impl.order_management.cost_analysis.cost_store import CostStore, normalize_part_no
from test_impl.order_management.cost_analysis.models import (
    INHOUSE_SUPPLIER_LABEL,
    LEGACY_PROCESS_ALIASES,
    PROCESS_BY_CODE,
    PROCESS_CODE_BY_NAME,
    is_inhouse_process,
)

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None  # type: ignore

_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


@dataclass
class _SheetGrid:
    """统一 openpyxl / xlrd 的 sheet 访问（行列均 1-based）。"""

    title: str
    max_row: int
    max_column: int
    _reader: Callable[[int, int], Any]

    def get_cell(self, row: int, col: int) -> Any:
        if row < 1 or col < 1:
            return None
        return self._reader(row, col)

# 表头标签 → 系统字段
_FIELD_LABELS: Dict[str, str] = {
    "客户": "customer_name",
    "产品名称": "product_name",
    "模具编号": "mold_no",
    "产品料号": "product_part_no",
    "文件编号": "product_part_no",
    "料号": "product_part_no",
    "模穴": "cavity",
    "产品单重": "unit_weight_g",
    "产品单重(G)": "unit_weight_g",
    "材质": "material",
    "机台&吨位": "machine_tonnage",
    "机台吨位": "machine_tonnage",
}

# Excel 工序写法 → 系统工序编号（员工确认 CL-0176）
_IMPORT_PROCESS_ALIASES: Dict[str, str] = {
    "压铸下料": "01",
    "压铸": "01",
    "埋轴": "01",
    "下料": "01",
    "精冲": "01",
    "解体精冲": "01",
    "精冲下料": "01",
    "手工下料": "01",
    "冲切下料": "01",
    "去毛刺": "02",
    "去毛边": "02",
    "去毛边攻牙": "02",
    "打磨去毛边": "02",
    "全检包装": "28",
    "全检": "28",
    "全检出货": "28",
    "全检贴膜": "28",
    "全检贴膜包装": "28",
    "包装": "34",
    "包装出货": "34",
    "振动研磨": "09",
    "研磨": "09",
    "皮膜钝化": "15",
    "皮膜": "15",
    "皮膜钝化（拉白）": "15",
    "钝化拉白": "15",
    "拉白": "15",
    "清洗拉白": "17",
    "清洗": "17",
    "铝挤": "32",
    "攻牙": "11",
    "钻孔攻牙倒角": "11",
}

# 一道 Excel 工序展开为多条系统工序（顺序保留）
_IMPORT_PROCESS_MULTI: Dict[str, List[str]] = {
    "振动研磨去毛边": ["02", "09"],
}

# 确认取消映射：跳过，不写入 BOM
_SKIP_PROCESS_NAMES = frozenset({"铆合弹片"})

_PLACEHOLDER_FIELD_VALUES = frozenset(
    {
        "文件编号/版次",
        "文件编号",
        "版次",
        "工序",
        "可加工厂商",
        "供应商",
        "模具&治具",
        "模具治具",
        "工装",
    }
)

_PROCESS_SEQ_RE = re.compile(r"^制程[:：]?\d+$", re.IGNORECASE)

_INHOUSE_SUPPLIER_WORDS = frozenset(
    {"厂内", "场内", "场内自制", "自制", "威可特", "昆山威可特"}
)

_SKIP_SHEET_NAMES = frozenset({"sheet", "模板", "说明", "目录"})
_PART_NO_REJECT_LABELS = frozenset({"数量", "单位", "备注", "序号", "制程", "工序"})


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_label(value: Any) -> str:
    s = _norm_text(value)
    s = s.replace("：", "").replace(":", "")
    s = s.replace("戶", "户")
    return re.sub(r"\s+", "", s)


_FIELD_LABEL_NORMS = frozenset(_norm_label(k) for k in _FIELD_LABELS)


def _is_placeholder_field(value: str) -> bool:
    raw = _norm_text(value)
    if not raw:
        return True
    if raw in _PLACEHOLDER_FIELD_VALUES:
        return True
    norm = _norm_label(raw)
    if norm in {_norm_label(x) for x in _PLACEHOLDER_FIELD_VALUES}:
        return True
    if "文件编号" in raw and "版次" in raw:
        return True
    return False


def _strip_process_name(raw: str) -> str:
    s = _norm_text(raw)
    s = re.sub(r"[（(].*?[）)]", "", s).strip()
    return s


def _fallback_from_sheet_title(title: str) -> str:
    t = _norm_text(title)
    if not t or t.casefold() in _SKIP_SHEET_NAMES:
        return ""
    if _is_placeholder_field(t):
        return ""
    return normalize_part_no(t)


def _cell_value(ws: _SheetGrid, row: int, col: int) -> str:
    try:
        return _norm_text(ws.get_cell(row, col))
    except Exception:
        return ""


def _value_beside_label(ws, row: int, col: int, *, max_offset: int = 4) -> str:
    for offset in range(1, max_offset + 1):
        val = _cell_value(ws, row, col + offset)
        if val and _norm_label(val) not in _FIELD_LABELS:
            return val
    return ""


def _is_field_value(value: str) -> bool:
    if not value or _is_placeholder_field(value):
        return False
    norm = _norm_label(value)
    if norm in _FIELD_LABEL_NORMS or norm in _PART_NO_REJECT_LABELS:
        return False
    return True


def _value_for_label(ws, row: int, col: int) -> str:
    """威可特 BOM 表：标签在上、数值在下的纵表 + 少量横表。"""
    candidates = [
        (row + 1, col),
        (row + 1, col - 1),
        (row, col + 1),
        (row, col + 2),
    ]
    for r, c in candidates:
        if r < 1 or c < 1:
            continue
        val = _cell_value(ws, r, c)
        if _is_field_value(val):
            return val
    return ""


def _parse_header_fields(ws) -> dict:
    fields: Dict[str, str] = {}
    max_row = min(ws.max_row or 8, 8)
    max_col = min(ws.max_column or 12, 12)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            label = _norm_label(_cell_value(ws, r, c))
            if not label:
                continue
            for key, field in _FIELD_LABELS.items():
                if label != _norm_label(key):
                    continue
                if field in fields:
                    continue
                val = _value_for_label(ws, r, c)
                if not val:
                    continue
                if field == "unit_weight_g":
                    if len(re.findall(r"[:：]\s*\d", val)) > 1:
                        fields[field] = val
                    else:
                        fields[field] = _parse_weight(val)
                else:
                    fields[field] = val
    return fields


def _parse_weight(raw: str) -> str:
    s = _norm_text(raw)
    if not s:
        return ""
    s = s.replace("克", "").replace("g", "").replace("G", "")
    s = s.replace(",", "")
    m = re.search(r"\d+(?:\.\d+)?", s)
    return m.group(0) if m else s


def _split_multi_values(raw: str) -> List[str]:
    text = _norm_text(raw)
    if not text:
        return []
    chunks = re.split(r"[\n\r/]+", text)
    return [c.strip() for c in chunks if c.strip()]


def _expand_short_product_names(names: List[str]) -> List[str]:
    if len(names) <= 1:
        return names
    first = names[0]
    prefix = re.sub(r"([A-Z])$", "", first)
    code_m = re.match(r"^([A-Z0-9.-]+)", first)
    code = code_m.group(1) if code_m else ""
    out = [first]
    for name in names[1:]:
        if name.startswith(prefix) or name.startswith(code):
            out.append(name)
        elif code and re.match(r"^[\u4e00-\u9fffA-Za-z]", name):
            out.append(f"{code}{name}")
        else:
            out.append(f"{prefix}{name}")
    return out


def _parse_multi_weights(raw: str, count: int) -> List[str]:
    text = _norm_text(raw)
    found = [_parse_weight(x) for x in re.findall(r"[:：]\s*(\d+(?:\.\d+)?)\s*[gG克]?", text)]
    found = [x for x in found if x]
    if len(found) >= count:
        return found[:count]
    single = _parse_weight(text)
    if single:
        return [single] * count
    return [""] * count


def _expand_multi_product_row(row: dict) -> List[dict]:
    parts = _split_multi_values(row.get("product_part_no", ""))
    if len(parts) <= 1:
        return [row]
    names = _expand_short_product_names(_split_multi_values(row.get("product_name", "")))
    while len(names) < len(parts):
        names.append(parts[len(names)])
    weights = _parse_multi_weights(row.get("unit_weight_g", ""), len(parts))
    expanded: List[dict] = []
    for idx, part in enumerate(parts):
        copy = dict(row)
        copy["product_part_no"] = normalize_part_no(part)
        copy["product_name"] = names[idx]
        if weights[idx]:
            copy["unit_weight_g"] = weights[idx]
        expanded.append(copy)
    return expanded


def _resolve_process_name(name: str) -> Tuple[Optional[str], Optional[str], List[str]]:
    """返回 (code, display_name, warnings)。"""
    raw = _strip_process_name(name)
    if not raw:
        return None, None, []
    warnings: List[str] = []
    if raw in _IMPORT_PROCESS_ALIASES:
        code = _IMPORT_PROCESS_ALIASES[raw]
        return code, PROCESS_BY_CODE[code], warnings
    if raw in PROCESS_CODE_BY_NAME:
        return PROCESS_CODE_BY_NAME[raw], raw, warnings
    if raw in PROCESS_BY_CODE:
        return raw, PROCESS_BY_CODE[raw], warnings
    if raw in LEGACY_PROCESS_ALIASES:
        code = LEGACY_PROCESS_ALIASES[raw]
        return code, PROCESS_BY_CODE[code], warnings
    for alias, code in _IMPORT_PROCESS_ALIASES.items():
        if alias in raw or raw in alias:
            return code, PROCESS_BY_CODE[code], [f"工序「{raw}」按「{PROCESS_BY_CODE[code]}」导入"]
    for pname, code in PROCESS_CODE_BY_NAME.items():
        if pname in raw or raw in pname:
            return code, pname, [f"工序「{raw}」按「{pname}」导入"]
    return None, None, [f"无法识别工序：{raw}"]


def _normalize_supplier(name: str, *, process_code: str) -> str:
    s = _norm_text(name)
    if is_inhouse_process(process_code):
        return ""
    if not s or s in _INHOUSE_SUPPLIER_WORDS:
        return INHOUSE_SUPPLIER_LABEL if not is_inhouse_process(process_code) else ""
    return s


def _find_process_header_row(ws) -> Optional[int]:
    max_row = min(ws.max_row or 40, 45)
    max_col = min(ws.max_column or 30, 40)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            text = _norm_label(_cell_value(ws, r, c))
            if text == "工序1" or text.startswith("工序1"):
                return r
    return None


def _parse_process_columns(ws, header_row: int) -> List[Tuple[int, int]]:
    """每道工序：(名称列, 供应商列)。"""
    max_col = min(ws.max_column or 30, 40)
    slots: List[Tuple[int, int]] = []
    for c in range(1, max_col + 1):
        text = _norm_label(_cell_value(ws, header_row, c))
        if re.match(r"^工序\d+$", text):
            name_col = c
            supplier_col = c + 1
            if _norm_label(_cell_value(ws, header_row, supplier_col)) != "供应商":
                supplier_col = c + 1
            slots.append((name_col, supplier_col))
    return slots


def _append_process(
    processes: List[dict],
    warnings: List[str],
    *,
    pname: str,
    supplier_raw: str,
) -> None:
    raw = _strip_process_name(pname)
    if not raw or raw in _SKIP_PROCESS_NAMES:
        return
    if raw in _IMPORT_PROCESS_MULTI:
        for code in _IMPORT_PROCESS_MULTI[raw]:
            display = PROCESS_BY_CODE[code]
            supplier = _normalize_supplier(supplier_raw, process_code=code)
            if not is_inhouse_process(code) and not supplier:
                warnings.append(f"外发工序「{display}」缺少供应商")
            processes.append(
                {
                    "code": code,
                    "name": display,
                    "supplier": supplier,
                    "price": "0",
                    "raw_name": pname,
                }
            )
        return
    code, display, w = _resolve_process_name(pname)
    warnings.extend(w)
    if not code:
        return
    supplier = _normalize_supplier(supplier_raw, process_code=code)
    if not is_inhouse_process(code) and not supplier:
        warnings.append(f"外发工序「{display}」缺少供应商")
    processes.append(
        {
            "code": code,
            "name": display,
            "supplier": supplier,
            "price": "0",
            "raw_name": pname,
        }
    )


def _dedupe_processes(processes: List[dict]) -> List[dict]:
    """相邻重复工序合并；跨步重复保留（如去毛边+震研）。"""
    deduped: List[dict] = []
    for p in processes:
        if deduped and deduped[-1]["code"] == p["code"]:
            continue
        deduped.append(p)
    return deduped


def _parse_processes_horizontal(ws: _SheetGrid) -> Tuple[List[dict], List[str]]:
    header_row = _find_process_header_row(ws)
    if header_row is None:
        return [], []

    slots = _parse_process_columns(ws, header_row)
    if not slots:
        return [], []

    processes: List[dict] = []
    warnings: List[str] = []
    data_start = header_row + 1
    data_end = min(ws.max_row or data_start, data_start + 5)
    stop_words = ("组装配件", "包材", "备注")

    for r in range(data_start, data_end + 1):
        first = _cell_value(ws, r, 1)
        if any(w in first for w in stop_words):
            break
        row_has_process = False
        for name_col, supplier_col in slots:
            pname = _cell_value(ws, r, name_col)
            if not pname or _norm_label(pname) in ("工序1", "供应商", "工装"):
                continue
            if re.match(r"^工序\d+$", _norm_label(pname)):
                continue
            row_has_process = True
            _append_process(
                processes,
                warnings,
                pname=pname,
                supplier_raw=_cell_value(ws, r, supplier_col),
            )
        if row_has_process:
            break

    return _dedupe_processes(processes), warnings


def _find_vertical_process_columns(ws: _SheetGrid) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    max_row = min(ws.max_row or 60, 80)
    max_col = min(ws.max_column or 20, 25)
    process_col: Optional[int] = None
    supplier_col: Optional[int] = None
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            label = _norm_label(_cell_value(ws, r, c))
            if label == "工序":
                process_col = c
            elif label in ("可加工厂商", "供应商"):
                supplier_col = c
    seq_col: Optional[int] = None
    for r in range(1, max_row + 1):
        for c in range(1, min(5, max_col + 1)):
            if _PROCESS_SEQ_RE.match(_norm_label(_cell_value(ws, r, c))):
                seq_col = c
                break
        if seq_col:
            break
    if process_col is None and seq_col is not None:
        process_col = seq_col + 1
    if supplier_col is None and process_col is not None:
        supplier_col = process_col + 1
    return seq_col, process_col, supplier_col


def _parse_processes_vertical(ws: _SheetGrid) -> Tuple[List[dict], List[str]]:
    seq_col, process_col, supplier_col = _find_vertical_process_columns(ws)
    if process_col is None:
        return [], ["未找到产品制程表"]

    processes: List[dict] = []
    warnings: List[str] = []
    max_row = min(ws.max_row or 60, 80)

    for r in range(1, max_row + 1):
        if seq_col is not None:
            if not _PROCESS_SEQ_RE.match(_norm_label(_cell_value(ws, r, seq_col))):
                continue
        else:
            pname_probe = _cell_value(ws, r, process_col)
            if not pname_probe or _is_placeholder_field(pname_probe):
                continue
            if _norm_label(pname_probe) == "工序":
                continue

        pname = _cell_value(ws, r, process_col)
        if not pname or _is_placeholder_field(pname):
            continue
        if _PROCESS_SEQ_RE.match(_norm_label(pname)):
            continue

        supplier_raw = _cell_value(ws, r, supplier_col) if supplier_col else ""
        _append_process(processes, warnings, pname=pname, supplier_raw=supplier_raw)

    if not processes:
        return [], ["未解析到任何工序"]
    return _dedupe_processes(processes), warnings


def _parse_processes(ws: _SheetGrid) -> Tuple[List[dict], List[str]]:
    processes, warnings = _parse_processes_horizontal(ws)
    if processes:
        return processes, warnings
    return _parse_processes_vertical(ws)


def _parse_sheet(ws: _SheetGrid) -> Optional[dict]:
    fields = _parse_header_fields(ws)
    sheet_part = _fallback_from_sheet_title(ws.title)
    part_raw = fields.get("product_part_no", "")
    if (
        _is_placeholder_field(part_raw)
        or not _norm_text(part_raw)
        or _norm_label(part_raw) in _PART_NO_REJECT_LABELS
    ):
        if sheet_part:
            fields["product_part_no"] = sheet_part
    product_raw = fields.get("product_name", "")
    if _is_placeholder_field(product_raw) or not _norm_text(product_raw):
        if sheet_part or _norm_text(ws.title):
            fields["product_name"] = _norm_text(ws.title)

    has_identity = bool(
        _norm_text(fields.get("product_part_no"))
        or _norm_text(fields.get("customer_name"))
        or sheet_part
    )
    if not has_identity:
        return None

    processes, proc_warnings = _parse_processes(ws)
    part_no = normalize_part_no(fields.get("product_part_no", "") or sheet_part)
    warnings = list(proc_warnings)
    row = {
        "sheet_name": _norm_text(ws.title),
        "customer_name": fields.get("customer_name", ""),
        "product_name": fields.get("product_name", ""),
        "mold_no": fields.get("mold_no", ""),
        "product_part_no": part_no,
        "cavity": fields.get("cavity", ""),
        "unit_weight_g": fields.get("unit_weight_g", ""),
        "material": fields.get("material", ""),
        "machine_tonnage": fields.get("machine_tonnage", ""),
        "processes": processes,
        "warnings": warnings,
    }
    return row


def _guess_format(file_bytes: bytes, filename: str = "") -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "xls":
        return "xls"
    if ext in ("xlsx", "xlsm"):
        return "xlsx"
    if file_bytes[:8] == _XLS_MAGIC:
        return "xls"
    if file_bytes[:2] == b"PK":
        return "xlsx"
    return "xlsx"


def _load_xlsx_sheets(file_bytes: bytes) -> List[_SheetGrid]:
    if load_workbook is None:
        raise ValueError("服务器未安装 openpyxl，无法解析 .xlsx")
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
    sheets: List[_SheetGrid] = []
    try:
        for name in wb.sheetnames:
            ws = wb[name]

            def _reader(r: int, c: int, _ws=ws) -> Any:
                return _ws.cell(row=r, column=c).value

            sheets.append(
                _SheetGrid(
                    title=str(name),
                    max_row=int(ws.max_row or 0),
                    max_column=int(ws.max_column or 0),
                    _reader=_reader,
                )
            )
    finally:
        wb.close()
    return sheets


def _load_xls_sheets(file_bytes: bytes) -> List[_SheetGrid]:
    if xlrd is None:
        raise ValueError("读取 .xls 需要安装 xlrd：pip install \"xlrd==1.2.0\"")
    version = getattr(xlrd, "__VERSION__", "") or getattr(xlrd, "__version__", "")
    if str(version).startswith("2."):
        raise ValueError(
            "当前 xlrd 2.x 不支持 .xls，请在服务器执行：pip install \"xlrd==1.2.0\""
        )
    book = xlrd.open_workbook(file_contents=file_bytes)
    sheets: List[_SheetGrid] = []
    for sheet in book.sheets():

        def _reader(r: int, c: int, _sh=sheet) -> Any:
            ri, ci = r - 1, c - 1
            if ri < 0 or ci < 0 or ri >= _sh.nrows or ci >= _sh.ncols:
                return None
            return _sh.cell_value(ri, ci)

        sheets.append(
            _SheetGrid(
                title=str(sheet.name),
                max_row=int(sheet.nrows),
                max_column=int(sheet.ncols),
                _reader=_reader,
            )
        )
    return sheets


def _load_workbook_sheets(file_bytes: bytes, *, filename: str = "") -> List[_SheetGrid]:
    fmt = _guess_format(file_bytes, filename)
    if fmt == "xls":
        return _load_xls_sheets(file_bytes)
    return _load_xlsx_sheets(file_bytes)


def parse_bom_workbook(file_bytes: bytes, *, filename: str = "") -> List[dict]:
    sheets = _load_workbook_sheets(file_bytes, filename=filename)
    rows: List[dict] = []
    for ws in sheets:
        if _norm_text(ws.title).casefold() in _SKIP_SHEET_NAMES:
            continue
        parsed = _parse_sheet(ws)
        if parsed:
            rows.extend(_expand_multi_product_row(parsed))
    return rows


def build_import_payload(parsed: dict) -> dict:
    customer = str(parsed.get("customer_name") or "").strip()
    if customer in _FILENAME_CUSTOMER_ALIASES:
        customer = _FILENAME_CUSTOMER_ALIASES[customer]
    process_prices: Dict[str, Any] = {}
    process_suppliers: Dict[str, str] = {}
    process_order: List[str] = []
    for p in parsed.get("processes") or []:
        code = str(p.get("code") or "").strip()
        if not code:
            continue
        process_order.append(code)
        process_prices[code] = "0"
        supplier = str(p.get("supplier") or "").strip()
        if supplier and not is_inhouse_process(code):
            process_suppliers[code] = supplier

    return {
        "customer_name": customer,
        "product_name": parsed.get("product_name", ""),
        "mold_no": parsed.get("mold_no", "") or "待补",
        "product_part_no": parsed.get("product_part_no", ""),
        "cavity": parsed.get("cavity", "") or "1*1",
        "unit_weight_g": parsed.get("unit_weight_g", "") or "0",
        "material": parsed.get("material", "") or "ADC12",
        "machine_tonnage": parsed.get("machine_tonnage", "") or "待补",
        "material_unit_price": "0",
        "process_prices": process_prices,
        "process_suppliers": process_suppliers,
        "process_order": process_order,
    }


def _tier_for_row(parsed: dict, store: CostStore) -> Tuple[str, List[str]]:
    issues: List[str] = list(parsed.get("warnings") or [])
    part = normalize_part_no(parsed.get("product_part_no", ""))
    customer = _norm_text(parsed.get("customer_name"))
    product_name = _norm_text(parsed.get("product_name"))

    if not customer:
        issues.append("缺少客户")
    if not part:
        issues.append("缺少产品料号")
    if not product_name:
        issues.append("缺少产品名称")
    if not parsed.get("processes"):
        issues.append("未解析到任何工序")

    for key, label in (
        ("mold_no", "模具编号"),
        ("unit_weight_g", "产品单重"),
        ("machine_tonnage", "机台吨位"),
    ):
        if not _norm_text(parsed.get(key)):
            issues.append(f"缺少{label}")

    if part:
        binding = store.get_part_binding(part)
        if binding and binding.get("customer_name"):
            issues.append(
                f"料号已存在（客户：{binding['customer_name']}），导入时将按料号覆盖"
            )

    blocking = [
        i
        for i in issues
        if i.startswith("缺少客户")
        or i.startswith("缺少产品料号")
        or i.startswith("缺少产品名称")
        or i.startswith("未解析")
        or i.startswith("无法识别工序")
        or i.startswith("未找到客户")
        or i.startswith("无法从文件名")
        or i.startswith("文件名「")
        or i.startswith("系统尚无客户")
    ]
    if blocking:
        return "blocked", issues
    if issues:
        return "pending", issues
    return "passed", issues


def _resolve_sheet_customer(raw: str, customer_names: List[str]) -> tuple[str, List[str]]:
    """Excel 客户简称 → 客商档案全称。"""
    hint = _norm_text(raw)
    warnings: List[str] = []
    if not hint:
        return "", warnings
    resolved, note = resolve_customer_from_hint(hint, customer_names)
    if resolved:
        if note:
            warnings.append(note)
        return resolved, warnings
    if note:
        warnings.append(note)
    return hint, warnings


def _apply_filename_customer(
    parsed_rows: List[dict],
    *,
    filename: str,
    customer_names: List[str],
) -> tuple[List[dict], dict]:
    hint = extract_customer_hint_from_filename(filename)
    resolved, customer_error = resolve_customer_from_hint(hint, customer_names)
    meta = {
        "customer_hint": hint,
        "customer_resolved": resolved or "",
        "customer_error": customer_error,
    }
    out: List[dict] = []
    for row in parsed_rows:
        copy = dict(row)
        warnings = list(copy.get("warnings") or [])
        sheet_customer = _norm_text(copy.get("customer_name"))
        if sheet_customer:
            copy["sheet_customer_raw"] = sheet_customer
            resolved_c, c_warnings = _resolve_sheet_customer(sheet_customer, customer_names)
            copy["customer_name"] = resolved_c or sheet_customer
            warnings.extend(c_warnings)
        elif resolved:
            copy["customer_name"] = resolved
        elif customer_error:
            warnings.append(customer_error)
        copy["warnings"] = warnings
        out.append(copy)

    row_customers = dedupe_customer_names(
        _norm_text(r.get("customer_name")) for r in out if _norm_text(r.get("customer_name"))
    )
    if row_customers:
        if len(row_customers) == 1:
            meta["customer_resolved"] = row_customers[0]
            meta["customer_error"] = ""
        else:
            joined = "、".join(row_customers[:5])
            extra = f" 等{len(row_customers)}个" if len(row_customers) > 5 else ""
            meta["customer_resolved"] = f"{joined}{extra}"
            meta["customer_error"] = ""
    return out, meta


def revalidate_preview_item(
    parsed: dict,
    customer_name: str,
    *,
    store: CostStore,
    fields: Optional[dict] = None,
) -> dict:
    """人工确认/修改字段后重新计算档位与 payload。"""
    row = dict(parsed or {})
    overrides = dict(fields or {})
    if customer_name:
        overrides.setdefault("customer_name", customer_name)
    for key in (
        "customer_name",
        "product_part_no",
        "product_name",
        "unit_weight_g",
        "mold_no",
        "machine_tonnage",
        "material",
        "cavity",
    ):
        if key not in overrides:
            continue
        val = str(overrides.get(key) or "").strip()
        if val in ("—", "-"):
            val = ""
        row[key] = val
    confirmed = _norm_text(row.get("customer_name"))
    row["customer_name"] = confirmed
    row["warnings"] = [
        w
        for w in (row.get("warnings") or [])
        if not any(
            str(w).startswith(p)
            for p in (
                "未找到客户",
                "无法从文件名",
                "文件名「",
                "系统尚无客户",
            )
        )
    ]
    raw = _norm_text(row.get("sheet_customer_raw"))
    if confirmed and raw and raw != confirmed:
        row["warnings"].append(f"客户已人工确认为「{confirmed}」（表内原为「{raw}」）")
    tier, issues = _tier_for_row(row, store)
    return {
        "tier": tier,
        "issues": issues,
        "parsed": row,
        "payload": build_import_payload(row),
        "process_display": " → ".join(
            p.get("name") or p.get("code", "") for p in row.get("processes") or []
        ),
    }


def preview_import_rows(
    parsed_rows: List[dict],
    *,
    store: CostStore,
) -> List[dict]:
    previews: List[dict] = []
    for idx, row in enumerate(parsed_rows):
        tier, issues = _tier_for_row(row, store)
        previews.append(
            {
                "index": idx,
                "sheet_name": row.get("sheet_name", ""),
                "tier": tier,
                "issues": issues,
                "parsed": row,
                "payload": build_import_payload(row),
                "process_display": " → ".join(
                    p.get("name") or p.get("code", "") for p in row.get("processes") or []
                ),
            }
        )
    return previews


def preview_import_batch(
    parsed_rows: List[dict],
    *,
    store: CostStore,
    filename: str = "",
    customer_names: Optional[List[str]] = None,
) -> dict:
    rows, meta = _apply_filename_customer(
        parsed_rows,
        filename=filename,
        customer_names=customer_names or [],
    )
    items = preview_import_rows(rows, store=store)
    return {**meta, "items": items}
