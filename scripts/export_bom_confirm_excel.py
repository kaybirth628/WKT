#!/usr/bin/env python3
"""导出 BOM 工序确认 Excel（简洁版，供员工填写）。"""
from __future__ import annotations

import re
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("需要 openpyxl：pip install openpyxl") from exc

from scripts.audit_bom_excel import _collect_files, _load_customer_names, audit_all
from test_impl.order_management.cost_analysis.cost_store import CostStore
from test_impl.order_management.order_entry.line_store import default_db_path

IN_DIR = ROOT / "data" / "bom_import_audit"
OUT = IN_DIR / "确认-工序映射.xlsx"

# 模糊别名：同一组 Excel 写法共用一次确认（与截图一致）
ALIAS_GROUPS = [
    ("解体精冲 / 精冲下料", "压铸"),
    ("振动研磨 / 振动研磨去毛边", "震研 / 去毛边"),
    ("去毛边攻牙 / 打磨去毛边", "去毛边"),
    ("清洗", "超声波清洗"),
    ("铆合弹片", "铆合"),
    ("皮膜 / 皮膜钝化（拉白）", "皮模钝化"),
    ("攻牙 / 钻孔攻牙倒角", "钻孔攻牙"),
]

_UNKNOWN_RE = re.compile(r"无法识别工序：(.+)")


def _count_unknown(items) -> Counter[str]:
    counts: Counter[str] = Counter()
    for item in items:
        m = _UNKNOWN_RE.search(item.detail or "")
        if m:
            counts[m.group(1).strip()] += 1
    return counts


def _write_sheet(ws, title: str, headers: list[str], rows: list[list]) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1, value=title)
    t.font = Font(bold=True, size=12, name="Microsoft YaHei")
    t.alignment = Alignment(vertical="center")

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True, name="Microsoft YaHei")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(rows, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Microsoft YaHei")
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {1: 28, 2: 14, 3: 22}
    for col, w in widths.items():
        if col <= len(headers):
            ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A3"


def main() -> int:
    IN_DIR.mkdir(parents=True, exist_ok=True)
    files = _collect_files([IN_DIR])
    store = CostStore(db_path=default_db_path())
    try:
        items = audit_all(files, store, _load_customer_names()) if files else []
    finally:
        store._conn.close()

    unknown = _count_unknown(items)
    unknown_order = ["铝挤", "清洗拉白", "钝化拉白", "拉白"]
    unknown_rows = [[name, unknown.get(name, 0), ""] for name in unknown_order]

    alias_rows = [[excel, system, ""] for excel, system in ALIAS_GROUPS]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "未知工序"
    _write_sheet(
        ws1,
        "二、未知工序（需统一映射，共 4 种）",
        ["Excel 写法", "出现次数", "应映射为系统哪道工序？"],
        unknown_rows,
    )

    ws2 = wb.create_sheet("模糊工序别名")
    _write_sheet(
        ws2,
        "三、模糊工序别名（确认一次即可，全库通用）",
        ["Excel 写法", "系统当前按…导入", "是否正确？"],
        alias_rows,
    )

    wb.save(OUT)
    print(f"Wrote {OUT}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
